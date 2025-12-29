import streamlit as st
import pandas as pd
import os
import base64
import platform
import subprocess
import openpyxl
import re
import io 
import time 
import json
import threading
import queue
from datetime import datetime

# Conditional import for tkinter (not available on Streamlit Cloud/headless environments)
try:
    import tkinter as tk
    from tkinter import filedialog
    HAS_TKINTER = True
except ImportError:
    HAS_TKINTER = False
    tk = None
    filedialog = None 

# --- Library เสริม (streamlit-pdf-viewer) ---
try:
    from streamlit_pdf_viewer import pdf_viewer
    HAS_PDF_VIEWER = True
except ImportError:
    HAS_PDF_VIEWER = False

# --- Library สำหรับ PDF Text Search (PyMuPDF) ---
try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False
    # Warning จะแสดงเมื่อมีการใช้งานจริง (ใน find_text_bbox_in_pdf)

# --- CONFIGURATION (Cross-Platform) ---
import shutil

# ใช้ path แบบ relative กับตำแหน่งที่ app.py อยู่
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VENDOR_MASTER_PATH = os.path.join(SCRIPT_DIR, "Vendor_branch.xlsx")
CONFIG_FILE = "config.json"

def get_default_output_path():
    """Get default output path based on operating system"""
    if platform.system() == 'Windows':
        output_path = r"D:\Project\ocr\output"
        os.makedirs(output_path, exist_ok=True)
        return output_path
    else:
        # On Linux/Cloud: use script directory or /mount/src/view_ocr
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Try /mount/src/view_ocr/output (Streamlit Cloud)
        if os.path.exists("/mount/src/view_ocr"):
            output_path = "/mount/src/view_ocr/output"
            os.makedirs(output_path, exist_ok=True)
            return output_path
        # Otherwise use output folder in script directory
        output_path = os.path.join(script_dir, "output")
        os.makedirs(output_path, exist_ok=True)
        return output_path

def get_default_poppler_path():
    """Get Poppler path based on operating system"""
    system = platform.system()
    
    if system == 'Windows':
        possible_paths = [
            r"C:\poppler\Library\bin",
            r"C:\poppler\bin",
            r"C:\Program Files\poppler\bin",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        return None
    else:
        # Linux/macOS: use system PATH
        return None

def get_default_tesseract_path():
    """Get Tesseract path based on operating system"""
    system = platform.system()
    
    if system == 'Windows':
        possible_paths = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        return None
    else:
        # Linux/macOS: use system PATH (tesseract should be in PATH)
        if shutil.which("tesseract"):
            return shutil.which("tesseract")
        return None

# Get paths from environment variables or auto-detect
DEFAULT_OUTPUT_PATH = os.environ.get("OCR_OUTPUT_PATH", get_default_output_path())
POPPLER_PATH = os.environ.get("POPPLER_PATH", get_default_poppler_path())
TESSERACT_PATH = os.environ.get("TESSERACT_PATH", get_default_tesseract_path())

# 1. ตั้งค่าหน้าจอ
st.set_page_config(layout="wide", page_title="AI OCR & Document Editor", page_icon="📄")

# CSS ปรับแต่ง
st.markdown("""
<style>
    .block-container { padding-top: 3rem !important; padding-bottom: 2rem; }
    /* PDF Viewer column sticky - only apply to column 2 in Document Editor */
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-of-type(2) {
        position: sticky; top: 3.5rem; height: 92vh; overflow-y: auto; align-self: start;
    }
    h3, h4 { margin-top: 0 !important; margin-bottom: 10px !important; padding-top: 0 !important; }
    div[data-testid="stButton"] button, div[data-testid="stDownloadButton"] button {
        height: 2.4rem; padding-top: 0; padding-bottom: 0; margin-top: 0px;
    }
    div[data-testid="column"] { gap: 0.5rem; }
    
    /* ปุ่มลูกศร */
    div[data-testid="column"] button:contains("⬆️"), div[data-testid="column"] button:contains("⬇️") {
        padding: 0 5px !important;
    }
    
    /* Compact page selector */
    div[data-testid="stSelectbox"] {
        margin-top: 0 !important;
    }
    div[data-testid="stSelectbox"] label {
        font-size: 0.9rem !important;
        margin-bottom: 0.2rem !important;
    }
    
    /* Compact page selection UI */
    div[data-testid="stNumberInput"] {
        margin-top: 0 !important;
    }
    div[data-testid="stNumberInput"] label {
        font-size: 0.85rem !important;
        margin-bottom: 0.1rem !important;
    }
    div[data-testid="stNumberInput"] input {
        padding: 0.25rem 0.5rem !important;
        font-size: 0.9rem !important;
    }
    div[data-testid="stHorizontalBlock"] {
        gap: 0.5rem !important;
    }
    
    /* Run OCR button - bold and blue */
    div[data-testid="stButton"] button[key="run_ocr_btn"] {
        font-weight: bold !important;
        color: #1f77b4 !important;
    }
    
    /* Select Folder button - green background */
    div[data-testid="stButton"] button[key="select_folder_btn"] {
        background-color: #28a745 !important;
        border-color: #28a745 !important;
    }
    div[data-testid="stButton"] button[key="select_folder_btn"]:hover {
        background-color: #218838 !important;
        border-color: #1e7e34 !important;
    }
    
    /* Increase font size for Source/Output Folder Path labels */
    div[data-testid="stTextInput"][data-baseweb="input"] label p,
    div[data-testid="stTextInput"] label p {
        font-size: 1.07em !important;
    }
    
    /* Align Settings button with Page selectbox - using nth-child for 4th column */
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(4) {
        display: flex !important;
        align-items: flex-start !important;
    }
    div[data-testid="stHorizontalBlock"] > div[data-testid="column"]:nth-child(4) > div {
        padding-top: 1.5rem !important;
    }
    
    /* Fix file uploader drag and drop */
    div[data-testid="stFileUploader"] {
        width: 100% !important;
    }
    div[data-testid="stFileUploaderDropzone"] {
        border: 2px dashed #ccc !important;
        border-radius: 10px !important;
        padding: 2.5rem 2rem !important;
        min-height: 180px !important;
        text-align: center !important;
        cursor: pointer !important;
        transition: all 0.3s ease !important;
        pointer-events: auto !important;   /* ensure drop events are received */
        position: relative !important;
        z-index: 2 !important;
        width: 100% !important;
        box-sizing: border-box !important;
        display: block !important;
    }
    div[data-testid="stFileUploaderDropzone"]:hover {
        border-color: #1f77b4 !important;
        background-color: #f0f8ff !important;
    }
    div[data-testid="stFileUploaderDropzoneInstructions"] {
        color: #666 !important;
        font-size: 0.95rem !important;
    }
</style>
<script>
(function() {
    function styleRunOCRButton() {
        const buttons = document.querySelectorAll('button[data-testid="stBaseButton-secondary"]');
        buttons.forEach(btn => {
            if (btn.textContent.trim() === 'Run OCR') {
                btn.style.fontWeight = 'bold';
                btn.style.color = '#1f77b4';
            }
        });
    }
    
    // Run on page load
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', styleRunOCRButton);
    } else {
        styleRunOCRButton();
    }
    
    // Also run after Streamlit reruns
    const observer = new MutationObserver(styleRunOCRButton);
    observer.observe(document.body, { childList: true, subtree: true });
    
    // Style Select Folder button - green
    function styleSelectFolderButton() {
        const buttons = document.querySelectorAll('button[data-testid="stBaseButton-primary"]');
        buttons.forEach(btn => {
            if (btn.textContent.trim() === '📁') {
                btn.style.backgroundColor = '#28a745';
                btn.style.borderColor = '#28a745';
                btn.addEventListener('mouseenter', function() {
                    this.style.backgroundColor = '#218838';
                    this.style.borderColor = '#1e7e34';
                });
                btn.addEventListener('mouseleave', function() {
                    this.style.backgroundColor = '#28a745';
                    this.style.borderColor = '#28a745';
                });
            }
        });
    }
    
    // Run on page load
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', styleSelectFolderButton);
    } else {
        styleSelectFolderButton();
    }
    
    // Also run after Streamlit reruns
    const observer2 = new MutationObserver(styleSelectFolderButton);
    observer2.observe(document.body, { childList: true, subtree: true });
    
    // Increase font size for Source/Output Folder Path labels
    function increaseFolderPathLabelFont() {
        const labels = document.querySelectorAll('div[data-testid="stTextInput"] label p');
        labels.forEach(label => {
            const text = label.textContent || '';
            if (text.includes('Source Folder Path:') || text.includes('Output Folder Path:')) {
                label.style.fontSize = '1.07em';
            }
        });
    }
    
    // Run on page load
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', increaseFolderPathLabelFont);
    } else {
        increaseFolderPathLabelFont();
    }
    
    // Also run after Streamlit reruns
    const observer3 = new MutationObserver(increaseFolderPathLabelFont);
    observer3.observe(document.body, { childList: true, subtree: true });
    
    // Field Focus Listener for PDF Highlight
    function setupFieldFocusListener() {
        // ตั้งค่า event listener สำหรับ text input fields
        const textInputs = document.querySelectorAll('input[data-testid="stTextInput"]');
        textInputs.forEach(input => {
            // เพิ่ม event listener สำหรับ focus
            input.addEventListener('focus', function() {
                // เก็บข้อมูล field ที่ focus
                const fieldName = this.getAttribute('aria-label') || this.previousElementSibling?.textContent || '';
                const fieldValue = this.value;
                
                // ส่งข้อมูลไปยัง Streamlit (ใช้ custom event)
                window.dispatchEvent(new CustomEvent('field_focus', {
                    detail: {
                        field_name: fieldName,
                        field_value: fieldValue
                    }
                }));
            });
        });
    }
    
    // Run on page load
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', setupFieldFocusListener);
    } else {
        setupFieldFocusListener();
    }
    
    // Also run after Streamlit reruns
    const observer4 = new MutationObserver(setupFieldFocusListener);
    observer4.observe(document.body, { childList: true, subtree: true });
})();
</script>
""", unsafe_allow_html=True)

# --- Function Definitions (Config) ---
def load_config():
    """โหลด config จากไฟล์ config.json (รองรับทั้ง API_KEY และ POPPLER_PATH)"""
    config_path = os.path.join(os.getcwd(), CONFIG_FILE)
    default_cfg = {'API_KEY': '', 'POPPLER_PATH': None}
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # รองรับโครงสร้างเก่า (เก็บเป็น string API_KEY)
                if isinstance(config, str):
                    return {'API_KEY': config, 'POPPLER_PATH': None}
                if isinstance(config, dict):
                    return {
                        'API_KEY': config.get('API_KEY', ''),
                        'POPPLER_PATH': config.get('POPPLER_PATH') or None
                    }
        except Exception:
            # ใช้ค่า default ถ้าอ่านไม่ได้
            return default_cfg
    return default_cfg

def save_config(api_key, poppler_path=None):
    """บันทึก config ลงไฟล์ config.json"""
    config_path = os.path.join(os.getcwd(), CONFIG_FILE)
    try:
        config = {'API_KEY': api_key, 'POPPLER_PATH': poppler_path or None}
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

# --- Session State Initialization ---
if 'current_page' not in st.session_state:
    st.session_state.current_page = "Page 1: AI OCR Dashboard"

# Page 1 State
if 'ocr_source_folder' not in st.session_state:
    st.session_state.ocr_source_folder = None
if 'ocr_output_folder' not in st.session_state:
    st.session_state.ocr_output_folder = DEFAULT_OUTPUT_PATH
if 'ocr_file_list_refresh' not in st.session_state:
    st.session_state.ocr_file_list_refresh = 0
if 'ocr_page_config' not in st.session_state:
    st.session_state.ocr_page_config = "All"  # Default: All pages
if 'ocr_page_start' not in st.session_state:
    st.session_state.ocr_page_start = 1  # Default start page
if 'show_delete_confirm' not in st.session_state:
    st.session_state.show_delete_confirm = False
if 'show_settings' not in st.session_state:
    st.session_state.show_settings = False
if 'ocr_type' not in st.session_state:
    st.session_state.ocr_type = "API Typhoon"  # Default: API Typhoon
config_data = load_config()
if 'api_key' not in st.session_state:
    # โหลด API_KEY จาก config file
    st.session_state.api_key = config_data.get('API_KEY', '')
if 'poppler_path' not in st.session_state:
    st.session_state.poppler_path = config_data.get('POPPLER_PATH', POPPLER_PATH)

# อัปเดตตัวแปร POPPLER_PATH ให้สอดคล้องกับ session (ใช้ใน pdf2image)
if st.session_state.poppler_path:
    POPPLER_PATH = st.session_state.poppler_path

# Page 2 State (Document Editor)
if 'df_data' not in st.session_state: 
    st.session_state.df_data = None
if 'current_sheet' not in st.session_state: 
    st.session_state.current_sheet = None
if 'view_mode' not in st.session_state: 
    st.session_state.view_mode = 'list'
if 'selected_row_idx' not in st.session_state: 
    st.session_state.selected_row_idx = None
if 'uploaded_file_ref' not in st.session_state: 
    st.session_state.uploaded_file_ref = None
if 'base_folder_cache' not in st.session_state: 
    st.session_state.base_folder_cache = os.getcwd()
if 'loaded_file_path' not in st.session_state:
    st.session_state.loaded_file_path = None
if 'doc_editor_path' not in st.session_state:
    st.session_state.doc_editor_path = None
if 'vendor_master_df' not in st.session_state: 
    st.session_state.vendor_master_df = None
if 'data_version' not in st.session_state: 
    st.session_state.data_version = 0
if 'highlighted_field' not in st.session_state:
    st.session_state.highlighted_field = None  # {'field_name': str, 'field_value': str, 'row_idx': int}
if 'pdf_highlight_positions' not in st.session_state:
    st.session_state.pdf_highlight_positions = []

# --- Function Definitions (Shared) ---
def open_file_external(file_path):
    try:
        if platform.system() == 'Windows': 
            os.startfile(file_path)
        elif platform.system() == 'Darwin': 
            subprocess.call(('open', file_path))
        else: 
            subprocess.call(('xdg-open', file_path))
    except Exception as e: 
        st.error(f"Error opening file: {e}")

def is_headless_environment():
    """Check if running in a headless environment (no display)"""
    # Check for common headless indicators
    if os.environ.get('CODESPACES') == 'true':
        return True
    if os.environ.get('GITPOD_WORKSPACE_ID'):
        return True
    if not os.environ.get('DISPLAY') and platform.system() != 'Windows':
        return True
    # Check if running in Docker
    if os.path.exists('/.dockerenv'):
        return True
    return False

def select_folder_dialog(initial_dir=None):
    """Open folder selection dialog using tkinter - thread-safe version"""
    
    # Check if tkinter is available
    if not HAS_TKINTER:
        st.info("💡 **Tip:** ระบบนี้ทำงานบน Cloud/Server ไม่รองรับ Folder Dialog\n\n👉 กรุณาพิมพ์ path โดยตรงในช่อง input แทน\n\nตัวอย่าง: `/home/vscode/ocr/source`")
        return None
    
    # Check for headless environment first
    if is_headless_environment():
        st.info("💡 **Tip:** ระบบนี้ทำงานบน Cloud/Server ไม่รองรับ Folder Dialog\n\n👉 กรุณาพิมพ์ path โดยตรงในช่อง input แทน\n\nตัวอย่าง: `/home/vscode/ocr/source`")
        return None
    
    result_queue = queue.Queue()
    error_queue = queue.Queue()

    def _run_dialog():
        """Run tkinter dialog in a separate thread"""
        try:
            # Create Tk instance in this thread
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            # Open dialog
            folder_path = filedialog.askdirectory(
                title="Select Folder",
                initialdir=initial_dir if initial_dir and os.path.exists(initial_dir) else None
            )
            root.destroy()
            result_queue.put(folder_path if folder_path else None)
        except Exception as e:
            error_queue.put(str(e))
            result_queue.put(None)

    try:
        # Check if we're in the main thread (tkinter requirement)
        # If not, we need to use a workaround
        dialog_thread = threading.Thread(target=_run_dialog, daemon=True)
        dialog_thread.start()
        dialog_thread.join(timeout=30)

        if dialog_thread.is_alive():
            st.warning("⚠️ Folder selection dialog timed out. Please enter path manually.")
            return None

        # Check for errors
        try:
            error = error_queue.get_nowait()
            # Show friendly message instead of error for display issues
            if "display" in error.lower() or "DISPLAY" in error:
                st.info("💡 **Tip:** Folder picker ใช้ไม่ได้บน Server\n\n👉 กรุณาพิมพ์ path โดยตรงในช่อง input แทน")
            else:
                st.error(f"Error selecting folder: {error}")
            return None
        except queue.Empty:
            pass

        # Get result
        result = result_queue.get(timeout=1)
        return result

    except queue.Empty:
        st.warning("⚠️ Could not get folder selection result. Please enter path manually.")
        return None
    except Exception as e:
        error_msg = str(e)
        if "display" in error_msg.lower() or "DISPLAY" in error_msg:
            st.info("💡 **Tip:** Folder picker ใช้ไม่ได้บน Server\n\n👉 กรุณาพิมพ์ path โดยตรงในช่อง input แทน")
        else:
            st.error(f"Error selecting folder: {e}")
        return None

def get_files_in_folder(folder_path):
    """Get list of files in a folder"""
    if not folder_path or not os.path.exists(folder_path):
        return []
    try:
        files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        return sorted(files)
    except Exception as e:
        st.error(f"Error reading folder: {e}")
        return []

def save_excel_local(df, default_name, start_path, header=True):
    """
    ฟังก์ชั่นเปิดหน้าต่าง Save As ของ Windows (เพิ่ม parameter header) - thread-safe version

    พิเศษ:
    - ถ้า header=True (ใช้ใน Document Editor -> Save As)
      จะพยายาม **รักษา Hyperlink เดิม** (เช่นคอลัมน์ `Link PDF`) โดยอาศัยโครงสร้างจากไฟล์ที่โหลดเข้ามา (`uploaded_file_ref`)
    - ถ้า header=False (เช่น Gen SAP) จะใช้วิธีเดิม คือเขียนไฟล์ใหม่จาก DataFrame ตรง ๆ
    """
    result_queue = queue.Queue()
    error_queue = queue.Queue()

    def _run_save_dialog():
        """Run tkinter save dialog in a separate thread"""
        if not HAS_TKINTER:
            result_queue.put(None)
            error_queue.put("tkinter not available on this platform")
            return
        try:
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            file_path = filedialog.asksaveasfilename(
                title="บันทึกไฟล์ Excel (Save As)",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=start_path if start_path and os.path.exists(start_path) else None,
                initialfile=default_name
            )
            root.destroy()
            result_queue.put(file_path if file_path else None)
        except Exception as e:
            error_queue.put(str(e))
            result_queue.put(None)

    try:
        # 1) เปิด Save As dialog (หรือใช้ default path ถ้าไม่มี tkinter)
        if HAS_TKINTER:
            dialog_thread = threading.Thread(target=_run_save_dialog, daemon=True)
            dialog_thread.start()
            dialog_thread.join(timeout=30)  # รอสูงสุด 30 วินาที

            if dialog_thread.is_alive():
                return False, "Save dialog timed out. Please try again."

            # ตรวจสอบ error
            try:
                error = error_queue.get_nowait()
                return False, f"Error: {error}"
            except queue.Empty:
                pass

            # ดึง path ที่ผู้ใช้เลือก
            file_path = result_queue.get(timeout=1)
            if not file_path:
                return False, "Cancelled"
        else:
            # On Streamlit Cloud, save directly to default path (skip dialog)
            file_path = os.path.join(start_path if start_path and os.path.exists(start_path) else os.getcwd(), default_name)

        # 2) ถ้าเป็นงาน Gen SAP (header=False) -> ใช้วิธีเดิม
        if not header:
            try:
                df_save = df.drop(columns=["_chk"], errors='ignore')
            except Exception:
                df_save = df
            df_save.to_excel(file_path, index=False, header=header, engine="openpyxl")
            return True, file_path

        # 3) กรณี Document Editor (header=True) -> พยายามรักษาโครงสร้างเดิม + Hyperlink
        try:
            from openpyxl import load_workbook
        except ImportError:
            # ถ้าไม่มี openpyxl ให้ fallback เป็นวิธีเดิม
            try:
                df_save = df.drop(columns=["_chk"], errors='ignore')
            except Exception:
                df_save = df
            df_save.to_excel(file_path, index=False, header=header, engine="openpyxl")
            return True, file_path

        # ใช้ไฟล์ต้นฉบับจาก session ถ้ามี (เพื่อรักษา Hyperlink)
        orig_file = None
        sheet_name = None
        if 'uploaded_file_ref' in st.session_state and st.session_state.uploaded_file_ref is not None:
            orig_file = st.session_state.uploaded_file_ref
        elif 'loaded_file_path' in st.session_state and st.session_state.loaded_file_path is not None:
            orig_file = st.session_state.loaded_file_path
        if 'current_sheet' in st.session_state and st.session_state.current_sheet is not None:
            sheet_name = st.session_state.current_sheet

        if orig_file is not None:
            # โหลด workbook เดิม (ที่ยังมี Hyperlink)
            try:
                if hasattr(orig_file, 'seek'):
                    orig_file.seek(0)
                    wb = load_workbook(orig_file, data_only=False)
                else:
                    # It's a file path
                    wb = load_workbook(orig_file, data_only=False)
            except Exception as e:
                wb = None
        else:
            wb = None

        if wb is None:
            # ถ้าโหลด workbook เดิมไม่ได้ ให้ใช้ to_excel ปกติ
            try:
                df_save = df.drop(columns=["_chk"], errors='ignore')
            except Exception:
                df_save = df
            df_save.to_excel(file_path, index=False, header=header, engine="openpyxl")
            return True, file_path

        # เลือก sheet ที่จะทำงาน
        if not sheet_name or sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        # สร้าง map: header text -> column index
        header_map = {}
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val is None:
                continue
            header_map[str(header_val)] = col_idx

        # ให้แน่ใจว่ามีคอลัมน์ครบตาม df (ยกเว้น _chk)
        for col_name in df.columns:
            if col_name == "_chk":
                continue
            if col_name not in header_map:
                max_col += 1
                ws.cell(row=1, column=max_col).value = col_name
                header_map[col_name] = max_col

        # ตรวจสอบว่าคอลัมน์ไหนเป็น Link PDF
        link_pdf_cols = []
        for col_name in df.columns:
            if col_name == "_chk":
                continue
            col_lower = str(col_name).lower()
            if "link" in col_lower and "pdf" in col_lower:
                link_pdf_cols.append(col_name)

        # เขียนค่าจาก df ลงใน worksheet
        for row_idx, (_, row_series) in enumerate(df.iterrows(), start=2):
            for col_name, value in row_series.items():
                if col_name == "_chk":
                    continue
                col_idx = header_map.get(col_name)
                if not col_idx:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # จัดการคอลัมน์ Link PDF แบบพิเศษ
                if col_name in link_pdf_cols:
                    # ตรวจสอบว่ามี hyperlink หรือ formula เดิมอยู่หรือไม่
                    has_existing_hyperlink = cell.hyperlink is not None
                    has_existing_formula = cell.data_type == 'f' and cell.value and str(cell.value).strip().upper().startswith("=HYPERLINK")
                    
                    # แปลงค่าเป็น string
                    if isinstance(value, float) and pd.isna(value):
                        value_str = ""
                    else:
                        value_str = str(value).strip() if value else ""
                    
                    # ถ้าค่าเป็น None/nan/empty และมี hyperlink/formula เดิมอยู่ -> ไม่ต้องทำอะไร (รักษา hyperlink เดิม)
                    if not value_str or value_str.lower() in ['none', 'nan', '']:
                        if has_existing_hyperlink or has_existing_formula:
                            # ไม่ต้องเขียนทับ ให้รักษา hyperlink/formula เดิมไว้
                            continue
                        else:
                            # ไม่มี hyperlink และไม่มีค่า -> ตั้งเป็นค่าว่าง
                            cell.value = ""
                            continue
                    
                    # ถ้ามีค่าใหม่ -> สร้าง hyperlink ใหม่
                    # ตรวจสอบว่าเป็น HYPERLINK formula หรือไม่
                    if value_str.upper().startswith("=HYPERLINK"):
                        # ถ้าเป็น formula ให้ตั้งเป็น formula
                        cell.value = value_str
                        # ลบ hyperlink object ถ้ามี (เพราะใช้ formula แทน)
                        if cell.hyperlink:
                            cell.hyperlink = None
                    else:
                        # ถ้าเป็น path ธรรมดา ให้สร้าง hyperlink
                        # ลองหา path จาก value_str
                        path_to_use = value_str
                        
                        # ถ้าเป็น relative path ให้ resolve
                        if not os.path.isabs(path_to_use):
                            base = st.session_state.get('base_folder_cache', os.getcwd())
                            path_to_use = os.path.join(base, path_to_use)
                            path_to_use = os.path.normpath(path_to_use)
                        
                        # สร้าง hyperlink
                        try:
                            cell.hyperlink = path_to_use
                            cell.style = "Hyperlink"
                            # ตั้ง display text เป็นชื่อไฟล์
                            display_text = os.path.basename(path_to_use) if os.path.basename(path_to_use) else value_str
                            cell.value = display_text
                            # ลบ formula ถ้ามี (เพราะใช้ hyperlink object แทน)
                            if cell.data_type == 'f':
                                cell.value = display_text
                        except Exception:
                            # ถ้าสร้าง hyperlink ไม่ได้ ให้ตั้งเป็นค่า text ธรรมดา
                            cell.value = value_str
                else:
                    # คอลัมน์อื่น -> เขียนค่าปกติ
                    if isinstance(value, float) and pd.isna(value):
                        cell.value = ""
                    else:
                        cell.value = value

        # บันทึก workbook ไปยังไฟล์ใหม่
        wb.save(file_path)
        return True, file_path

    except queue.Empty:
        return False, "Could not get save dialog result. Please try again."
    except Exception as e:
        return False, f"Error saving file: {e}"

def save_txt_local(df, default_name, start_path, delimiter='\t'):
    """
    ฟังก์ชั่นเปิดหน้าต่าง Save As สำหรับบันทึกไฟล์ Text (.txt) - thread-safe version
    
    Args:
        df: DataFrame ที่จะบันทึก
        default_name: ชื่อไฟล์เริ่มต้น
        start_path: path เริ่มต้นสำหรับ dialog
        delimiter: ตัวคั่นระหว่างคอลัมน์ (default: tab '\t')
    """
    result_queue = queue.Queue()
    error_queue = queue.Queue()

    def _run_save_dialog():
        """Run tkinter save dialog in a separate thread"""
        if not HAS_TKINTER:
            result_queue.put(None)
            error_queue.put("tkinter not available on this platform")
            return
        try:
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            file_path = filedialog.asksaveasfilename(
                title="บันทึกไฟล์ Text (Save As)",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                initialdir=start_path if start_path and os.path.exists(start_path) else None,
                initialfile=default_name
            )
            root.destroy()
            result_queue.put(file_path if file_path else None)
        except Exception as e:
            error_queue.put(str(e))
            result_queue.put(None)

    # Check if tkinter is available
    if not HAS_TKINTER:
        # On Streamlit Cloud, save directly to default path (skip dialog)
        file_path = os.path.join(start_path if start_path and os.path.exists(start_path) else os.getcwd(), default_name)
    else:
        try:
            # 1) เปิด Save As dialog
            dialog_thread = threading.Thread(target=_run_save_dialog, daemon=True)
            dialog_thread.start()
            dialog_thread.join(timeout=30)  # รอสูงสุด 30 วินาที

            if dialog_thread.is_alive():
                return False, "Save dialog timed out. Please try again."

            # ตรวจสอบ error
            try:
                error = error_queue.get_nowait()
                return False, f"Error: {error}"
            except queue.Empty:
                pass

            # ดึง path ที่ผู้ใช้เลือก
            file_path = result_queue.get(timeout=1)
            if not file_path:
                return False, "Cancelled"
        except queue.Empty:
            return False, "Could not get save dialog result. Please try again."
        except Exception as e:
            return False, f"Error: {e}"

    # 2) บันทึก DataFrame เป็นไฟล์ .txt
    try:
        # ใช้ delimiter ที่กำหนด (default: tab)
        df.to_csv(file_path, sep=delimiter, index=False, header=False, encoding='utf-8')
        return True, file_path
    except Exception as e:
        return False, f"Error saving text file: {e}"

def find_text_bbox_in_pdf(pdf_path, search_text, page_num, field_name=None):
    """
    หา bounding box ของ text ใน PDF ด้วย Tesseract OCR
    Args:
        pdf_path: path to PDF file
        search_text: text to search for
        page_num: page number (1-indexed)
        field_name: name of the field (e.g., "Description") - used to adjust bounding box size
    Returns: list of dicts with 'x0', 'y0', 'x1', 'y1', 'page', 'text', 'method', 'page_width', 'page_height'
    """
    positions = []
    
    if not search_text or not str(search_text).strip():
        return positions
    
    # Clean search text
    search_str = str(search_text).strip()
    if search_str.isdigit() or re.match(r'^[\d\-\s]+$', search_str):
        clean_search = re.sub(r'\D', '', search_str)
    else:
        clean_search = search_str
    
    if not clean_search:
        return positions
    
    # สำหรับ Description field: ใช้ bounding box ที่กำหนดไว้เลย (ไม่ต้องรัน OCR)
    if field_name and 'description' in str(field_name).lower():
        # หา page dimensions จาก PDF
        if HAS_PYMUPDF:
            try:
                doc = fitz.open(pdf_path)
                if page_num > 0 and page_num <= len(doc):
                    page = doc[page_num - 1]
                    page_rect = page.rect
                    page_width = page_rect.width
                    page_height = page_rect.height
                    doc.close()
                    
                    # กำหนด bounding box สำหรับ Description
                    new_x0 = 40.0
                    original_width_ratio = 0.60  # 60% ของความกว้างหน้า
                    new_x1 = new_x0 + (page_width * original_width_ratio)
                    new_y0 = 290.0
                    line_height = 20
                    new_y1 = new_y0 + (line_height * 5.5)  # 110 points
                    
                    # จำกัดให้อยู่ในขอบเขตของหน้า
                    new_x0 = max(0, min(new_x0, page_width - 100))
                    new_x1 = max(new_x0 + 100, min(new_x1, page_width))
                    new_y1 = max(new_y0 + 20, min(new_y1, page_height))
                    
                    positions.append({
                        'x0': float(new_x0),
                        'y0': float(new_y0),
                        'x1': float(new_x1),
                        'y1': float(new_y1),
                        'page': page_num,
                        'text': search_str,
                        'method': 'description_fixed',
                        'page_width': float(page_width),
                        'page_height': float(page_height)
                    })
                    return positions
            except Exception as e:
                print(f"[DEBUG] Error getting page dimensions for Description: {e}")
                # Fall through to OCR methods if error
    
    # สำหรับ Sales Promotion field: ใช้ bounding box ที่กำหนดไว้เลย (ไม่ต้องรัน OCR)
    if field_name and ('sales' in str(field_name).lower() and 'promotion' in str(field_name).lower()):
        # หา page dimensions จาก PDF
        if HAS_PYMUPDF:
            try:
                doc = fitz.open(pdf_path)
                if page_num > 0 and page_num <= len(doc):
                    page = doc[page_num - 1]
                    page_rect = page.rect
                    page_width = page_rect.width
                    page_height = page_rect.height
                    doc.close()
                    
                    # กำหนด bounding box สำหรับ Sales Promotion
                    # คำนวณขนาดเดิม (หลังจากลด 50% แล้ว)
                    base_width = 156.0  # width หลังจากลด 50%
                    base_height = 50.0  # height หลังจากลด 50%
                    
                    # ขยายขนาดขึ้น 30%
                    new_width = base_width * 1.3  # 156 * 1.3 = 202.8
                    new_height = base_height * 1.3  # 50 * 1.3 = 65
                    
                    # x0 = 20 (เลื่อนมาทางซ้าย)
                    new_x0 = 20.0
                    new_x1 = new_x0 + new_width  # 20 + 202.8 = 222.8
                    
                    # y0 = 510 (คงเดิม)
                    new_y0 = 510.0
                    new_y1 = new_y0 + new_height  # 510 + 65 = 575
                    
                    # จำกัดให้อยู่ในขอบเขตของหน้า
                    new_x0 = max(0, min(new_x0, page_width - 10))
                    new_x1 = max(new_x0 + 10, min(new_x1, page_width))
                    new_y0 = max(0, min(new_y0, page_height - 10))
                    new_y1 = max(new_y0 + 10, min(new_y1, page_height))
                    
                    positions.append({
                        'x0': float(new_x0),
                        'y0': float(new_y0),
                        'x1': float(new_x1),
                        'y1': float(new_y1),
                        'page': page_num,
                        'text': search_str,
                        'method': 'sales_promotion_fixed',
                        'page_width': float(page_width),
                        'page_height': float(page_height)
                    })
                    return positions
            except Exception as e:
                print(f"[DEBUG] Error getting page dimensions for Sales Promotion: {e}")
                # Fall through to OCR methods if error
    
    # วิธีที่ 1: ใช้ Tesseract OCR เพื่อหา bounding box ที่แม่นยำ
    try:
        import pytesseract
        from pdf2image import convert_from_path
        from PIL import Image
        
        # Set Tesseract path
        if TESSERACT_PATH and os.path.exists(TESSERACT_PATH):
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        
        # Convert PDF page to image
        images = convert_from_path(
            pdf_path,
            first_page=page_num,
            last_page=page_num,
            dpi=150,
            poppler_path=POPPLER_PATH if POPPLER_PATH else None
        )
        
        if not images:
            raise Exception("Failed to convert PDF to image")
        
        img = images[0]
        img_width, img_height = img.size
        
        # Run Tesseract OCR with bounding box data
        # Output format: TSV with columns: level, page_num, block_num, par_num, line_num, word_num, left, top, width, height, conf, text
        ocr_data = pytesseract.image_to_data(img, lang='tha+eng', output_type=pytesseract.Output.DICT)
        
        # สำหรับข้อความยาวๆ (multi-word) ใช้วิธีค้นหาคำหลายคำและรวม bounding boxes
        # แต่สำหรับคำเดียวหรือข้อความสั้นๆ ใช้วิธีเดิม
        search_words = clean_search.split()
        is_multi_word = len(search_words) > 1
        
        if is_multi_word and len(clean_search) > 10:
            # Multi-word search: หาคำต่างๆ ใน OCR results และรวม bounding boxes
            # สร้าง list ของ words พร้อม indices
            word_boxes = []
            for i, text in enumerate(ocr_data['text']):
                if not text.strip():
                    continue
                ocr_text_clean = text.strip()
                conf = ocr_data['conf'][i]
                if conf < 30:
                    continue
                
                word_boxes.append({
                    'text': ocr_text_clean,
                    'x': ocr_data['left'][i],
                    'y': ocr_data['top'][i],
                    'w': ocr_data['width'][i],
                    'h': ocr_data['height'][i],
                    'conf': conf,
                    'line': ocr_data['line_num'][i],
                    'index': i
                })
            
            # ค้นหาลำดับคำที่ match กับ search_words
            # ใช้ sliding window เพื่อหาลำดับคำที่ติดกัน
            best_match = None
            best_score = 0
            
            for start_idx in range(len(word_boxes)):
                matched_words = []
                search_idx = 0
                
                for word_box_idx in range(start_idx, len(word_boxes)):
                    word_box = word_boxes[word_box_idx]
                    search_word = search_words[search_idx] if search_idx < len(search_words) else None
                    
                    if not search_word:
                        break
                    
                    # เปรียบเทียบคำ
                    if (search_word.lower() in word_box['text'].lower() or 
                        word_box['text'].lower() in search_word.lower() or
                        (len(search_word) >= 3 and len(word_box['text']) >= 3 and 
                         search_word[:3].lower() == word_box['text'][:3].lower())):
                        matched_words.append(word_box)
                        search_idx += 1
                        
                        # ถ้า match ครบทุกคำ
                        if search_idx >= len(search_words):
                            # คำนวณ bounding box รวม
                            min_x = min(w['x'] for w in matched_words)
                            min_y = min(w['y'] for w in matched_words)
                            max_x = max(w['x'] + w['w'] for w in matched_words)
                            max_y = max(w['y'] + w['h'] for w in matched_words)
                            avg_conf = sum(w['conf'] for w in matched_words) / len(matched_words)
                            
                            match_score = 100 if search_idx == len(search_words) else int((search_idx / len(search_words)) * 100)
                            
                            if match_score > best_score:
                                best_score = match_score
                                best_match = {
                                    'x0': float(min_x),
                                    'y0': float(min_y),
                                    'x1': float(max_x),
                                    'y1': float(max_y),
                                    'page': page_num,
                                    'text': clean_search,
                                    'method': 'tesseract_ocr_multiword',
                                    'confidence': float(avg_conf),
                                    'match_score': match_score,
                                    'page_width': float(img_width),
                                    'page_height': float(img_height)
                                }
                                
                                # สำหรับ Description field: ขยาย bounding box
                                if field_name and 'description' in str(field_name).lower():
                                    new_x0 = 40.0  # x0 = 40 (คงที่)
                                    # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                                    original_width_ratio = 0.60  # 70% - 10% = 60%
                                    new_x1 = new_x0 + (img_width * original_width_ratio)
                                    line_height = 20
                                    # เลื่อนขึ้นไปขอบกระดาษบน (y0 = 0 หรือใกล้ 0)
                                    new_y0 = 290.0  # y0 = 290
                                    new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                                    new_x0 = max(0, min(new_x0, img_width - 100))
                                    new_x1 = max(new_x0 + 100, min(new_x1, img_width))
                                    new_y1 = max(new_y0 + 20, min(new_y1, img_height))
                                    best_match['x0'] = float(new_x0)
                                    best_match['x1'] = float(new_x1)
                                    best_match['y0'] = float(new_y0)
                                    best_match['y1'] = float(new_y1)
                            break
                    else:
                        # ถ้าไม่ match ให้เริ่มใหม่จากคำถัดไป
                        break
            
            if best_match and best_score >= 70:
                positions.append(best_match)
        
        else:
            # Single word หรือ short text search: ใช้วิธีเดิม
            for i, text in enumerate(ocr_data['text']):
                if not text.strip():
                    continue
                
                # Clean OCR text for comparison
                ocr_text_clean = text.strip()
                if ocr_text_clean.isdigit() or re.match(r'^[\d\-\s]+$', ocr_text_clean):
                    ocr_text_clean = re.sub(r'\D', '', ocr_text_clean)
                
                # Skip if too short
                if len(ocr_text_clean) < 3:
                    continue
                
                # Check if match - ใช้ exact match หรือ high similarity
                is_match = False
                match_score = 0
                
                # Exact match (best)
                if clean_search == ocr_text_clean:
                    is_match = True
                    match_score = 100
                # Full search text found in OCR text (good for long strings)
                elif len(clean_search) >= 8 and clean_search in ocr_text_clean:
                    is_match = True
                    match_score = 90
                # OCR text is complete substring of search (good)
                elif len(ocr_text_clean) >= 8 and ocr_text_clean in clean_search:
                    is_match = True
                    match_score = 85
                # High overlap for shorter strings (>70% match)
                elif len(clean_search) >= 5 and len(ocr_text_clean) >= 5:
                    # Calculate overlap ratio
                    if clean_search in ocr_text_clean or ocr_text_clean in clean_search:
                        overlap = min(len(clean_search), len(ocr_text_clean))
                        ratio = overlap / max(len(clean_search), len(ocr_text_clean))
                        if ratio >= 0.7:
                            is_match = True
                            match_score = int(ratio * 80)
                
                if not is_match:
                    continue
                
                # Get bounding box from Tesseract
                x = ocr_data['left'][i]
                y = ocr_data['top'][i]
                w = ocr_data['width'][i]
                h = ocr_data['height'][i]
                conf = ocr_data['conf'][i]
                
                # Skip low confidence results
                if conf < 30:
                    continue
                
                # Convert image coordinates to PDF coordinates (same scale for pdf2image)
                pos = {
                    'x0': float(x),
                    'y0': float(y),
                    'x1': float(x + w),
                    'y1': float(y + h),
                    'page': page_num,
                    'text': text.strip(),
                    'method': 'tesseract_ocr',
                    'confidence': float(conf),
                    'match_score': match_score,
                    'page_width': float(img_width),
                    'page_height': float(img_height)
                }
                
                # สำหรับ Description field: ขยาย bounding box
                if field_name and 'description' in str(field_name).lower():
                    new_x0 = 40.0  # x0 = 40 (คงที่)
                    # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                    original_width_ratio = 0.60  # 70% - 10% = 60%
                    new_x1 = new_x0 + (img_width * original_width_ratio)
                    line_height = 20
                    # y0 = 290 (เลื่อนลงมา)
                    new_y0 = 290.0  # y0 = 290
                    new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                    new_x0 = max(0, min(new_x0, img_width - 100))
                    new_x1 = max(new_x0 + 100, min(new_x1, img_width))
                    new_y1 = max(new_y0 + 20, min(new_y1, img_height))
                    pos['x0'] = float(new_x0)
                    pos['x1'] = float(new_x1)
                    pos['y0'] = float(new_y0)
                    pos['y1'] = float(new_y1)
                
                positions.append(pos)
        
        # ถ้าเจอแล้ว filter และเลือก best matches
        if positions:
            # Debug: แสดงข้อมูล matches ทั้งหมด
            print(f"\n[DEBUG] Tesseract found {len(positions)} potential matches for '{clean_search}':")
            for idx, p in enumerate(positions):
                print(f"  {idx+1}. Text='{p.get('text')}' | Match={p.get('match_score')}% | Conf={p.get('confidence'):.0f}% | Pos=({p.get('x0'):.0f},{p.get('y0'):.0f})")
            
            # Sort by match_score (descending) then confidence (descending)
            positions.sort(key=lambda p: (p.get('match_score', 0), p.get('confidence', 0)), reverse=True)
            
            # เอาแค่ top 3 matches หรือ matches ที่มี score เท่ากับ max score
            if len(positions) > 0:
                max_score = positions[0].get('match_score', 0)
                # เอา matches ที่มี score ใกล้เคียง max (ต่างกันไม่เกิน 10)
                filtered = [p for p in positions if p.get('match_score', 0) >= max_score - 10]
                # จำกัดไม่เกิน 3 boxes
                positions = filtered[:3]
                print(f"[DEBUG] Selected top {len(positions)} match(es) with max_score={max_score}")
                
                # สำหรับ Description field: ขยาย bounding box ก่อน return
                if field_name and 'description' in str(field_name).lower():
                    for pos in positions:
                        page_width = pos.get('page_width', 595)
                        page_height = pos.get('page_height', 842)
                        new_x0 = 40.0  # x0 = 40 (คงที่)
                        # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                        original_width_ratio = 0.60  # 70% - 10% = 60%
                        new_x1 = new_x0 + (page_width * original_width_ratio)
                        line_height = 20
                        # y0 = 290 (เลื่อนลงมา)
                        new_y0 = 290.0  # y0 = 290
                        new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                        new_x0 = max(0, min(new_x0, page_width - 100))
                        new_x1 = max(new_x0 + 100, min(new_x1, page_width))
                        new_y1 = max(new_y0 + 20, min(new_y1, page_height))
                        pos['x0'] = float(new_x0)
                        pos['x1'] = float(new_x1)
                        pos['y0'] = float(new_y0)
                        pos['y1'] = float(new_y1)
            
            return positions
            
    except ImportError:
        st.info("💡 Install Tesseract OCR for accurate text positioning: `pip install pytesseract` and install Tesseract binary")
    except Exception as e:
        print(f"[DEBUG] Tesseract OCR failed: {e}")
        # Fall through to other methods
    
    # วิธีที่ 1: ลองค้นหาใน PDF text layer ก่อน (สำหรับ PDF ที่มี text)
    if HAS_PYMUPDF:
        try:
            doc = fitz.open(pdf_path)
            if page_num > 0 and page_num <= len(doc):
                page = doc[page_num - 1]
                
                # ค้นหา text instances ใน PDF
                text_instances = page.search_for(clean_search)
                
                # ถ้าไม่เจอ ลองค้นหาแบบ partial
                if not text_instances and len(clean_search) > 3:
                    parts = re.findall(r'\d+', clean_search)
                    if parts:
                        for part in parts:
                            if len(part) >= 4:
                                partial_instances = page.search_for(part)
                                text_instances.extend(partial_instances)
                
                if text_instances:
                    # Get page dimensions
                    page_rect = page.rect
                    page_width = page_rect.width
                    page_height = page_rect.height
                    
                    for inst in text_instances:
                        pos = {
                            'x0': float(inst.x0),
                            'y0': float(inst.y0),
                            'x1': float(inst.x1),
                            'y1': float(inst.y1),
                            'page': page_num,
                            'text': clean_search,
                            'method': 'pdf_text_layer',
                            'page_width': float(page_width),
                            'page_height': float(page_height)
                        }
                        
                        # สำหรับ Description field: ขยาย bounding box ให้ครอบคลุมข้อความทั้งหมด
                        if field_name and 'description' in str(field_name).lower():
                            # x0 = 40 (คงที่)
                            new_x0 = 40.0
                            # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                            original_width_ratio = 0.60  # 70% - 10% = 60%
                            new_x1 = new_x0 + (page_width * original_width_ratio)
                            
                            # ขยายความสูง: ครอบคลุม 5.5 บรรทัด (ประมาณ 110 points)
                            # y0 = 290 (เลื่อนลงมา)
                            line_height = 20
                            new_y0 = 290.0  # y0 = 290
                            new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                            
                            # จำกัดให้อยู่ในขอบเขตของหน้า
                            new_x0 = max(0, min(new_x0, page_width - 100))
                            new_x1 = max(new_x0 + 100, min(new_x1, page_width))
                            new_y1 = max(new_y0 + 20, min(new_y1, page_height))
                            
                            # อัพเดท bounding box
                            pos['x0'] = float(new_x0)
                            pos['x1'] = float(new_x1)
                            pos['y0'] = float(new_y0)
                            pos['y1'] = float(new_y1)
                        
                        positions.append(pos)
            
            doc.close()
            
            # ถ้าเจอแล้ว return
            if positions:
                return positions
        except Exception as e:
            # ถ้า error ในการอ่าน PDF text layer ให้ลองวิธีอื่น
            pass
    
    # วิธีที่ 2: สำหรับ PDF แบบ scan - อ่านจากไฟล์ .txt ที่ OCR สร้างไว้
    try:
        # หาไฟล์ .txt ที่เกี่ยวข้อง
        pdf_dir = os.path.dirname(pdf_path)
        pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
        txt_filename = f"{pdf_basename}_page{page_num}.txt"
        
        # ลองหาใน output folder ก่อน
        possible_txt_paths = [
            os.path.join(pdf_dir, txt_filename),  # ในโฟลเดอร์เดียวกับ PDF
            os.path.join(st.session_state.get('ocr_output_folder', DEFAULT_OUTPUT_PATH), txt_filename),  # ใน output folder
            os.path.join(DEFAULT_OUTPUT_PATH, txt_filename),  # ใน default output folder
        ]
        
        txt_path = None
        for path in possible_txt_paths:
            if os.path.exists(path):
                txt_path = path
                break
        
        if txt_path and os.path.exists(txt_path):
            # อ่านไฟล์ .txt
            with open(txt_path, 'r', encoding='utf-8') as f:
                ocr_text = f.read()
            
            # ทำความสะอาด search_text เพื่อเปรียบเทียบกับ OCR text (ลบ HTML tags)
            # แต่เก็บ original search_text ไว้สำหรับ highlight
            search_for_matching = clean_search
            # ลบ HTML tags จาก search text (ถ้ามี) เพื่อให้ match ได้ดีขึ้น
            search_clean_html = re.sub(r'<[^>]+>', ' ', search_for_matching)
            search_clean_html = re.sub(r'\s+', ' ', search_clean_html).strip()
            
            # สร้าง clean version ของ OCR text เพื่อค้นหา
            ocr_text_clean = re.sub(r'<br/?>', ' ', ocr_text)
            ocr_text_clean = re.sub(r'<[^>]+>', '', ocr_text_clean)
            ocr_text_clean = re.sub(r'\s+', ' ', ocr_text_clean)
            
            # ค้นหาโดยใช้ลำดับความสำคัญ:
            # 1. Exact match ใน clean text (ดีที่สุด)
            # 2. Exact match ใน original text
            # 3. Partial match ที่ครอบคลุมมากที่สุด (สำหรับข้อความยาวๆ)
            found_match = None
            used_pattern = None
            match_quality = 0  # 0=not found, 1=partial, 2=exact clean, 3=exact original
            
            # สำหรับ Sales Promotion - ตรวจสอบ keyword เพื่อหาเฉพาะบรรทัดที่ถูกต้อง
            is_sales_promotion = 'ค่าส่งเสริมการขาย' in search_clean_html or 'Sales Promotion' in search_for_matching
            
            # วิธีที่ 1: Exact match ใน clean text (ดีที่สุด - ไม่มี HTML tags)
            if search_clean_html:
                # สำหรับ Sales Promotion ให้หาเฉพาะบรรทัดที่มี "ค่าส่งเสริมการขาย" โดยไม่รวม "หมายเหตุ"
                if is_sales_promotion:
                    # แบ่งเป็นบรรทัดทั้ง clean และ original
                    ocr_lines_clean = ocr_text_clean.split('\n')
                    ocr_lines_orig = ocr_text.split('\n')
                    
                    for line_idx, line_clean in enumerate(ocr_lines_clean):
                        # ตรวจสอบว่าเป็นบรรทัดที่มี "ค่าส่งเสริมการขาย" และไม่มี "หมายเหตุ"
                        if 'ค่าส่งเสริมการขาย' in line_clean and 'หมายเหตุ' not in line_clean:
                            # ตรวจสอบว่า search text อยู่ในบรรทัดนี้
                            if search_clean_html in line_clean:
                                # หาในบรรทัดนี้
                                line_match = re.search(re.escape(search_clean_html), line_clean, re.IGNORECASE)
                                if line_match:
                                    # หาบรรทัดที่ตรงกันใน original text
                                    if line_idx < len(ocr_lines_orig):
                                        line_orig = ocr_lines_orig[line_idx]
                                        # ลบ HTML tags จาก original line เพื่อหา position
                                        line_orig_clean = re.sub(r'<[^>]+>', '', line_orig)
                                        line_orig_clean = re.sub(r'\s+', ' ', line_orig_clean)
                                        
                                        # หา position ใน original line
                                        orig_line_match = re.search(re.escape(search_clean_html), line_orig_clean, re.IGNORECASE)
                                        if orig_line_match:
                                            # คำนวณตำแหน่งใน original full text
                                            chars_before_line = sum(len(l) + 1 for l in ocr_lines_orig[:line_idx])
                                            match_start = chars_before_line + orig_line_match.start()
                                            match_end = chars_before_line + orig_line_match.end()
                                            
                                            from collections import namedtuple
                                            MatchObj = namedtuple('MatchObj', ['start', 'end', 'group'])
                                            found_match = MatchObj(
                                                start=match_start, 
                                                end=match_end, 
                                                group=lambda: search_clean_html
                                            )
                                            match_quality = 3
                                            break
                else:
                    # สำหรับ fields อื่นๆ ใช้วิธีเดิม
                    exact_match_clean = re.search(re.escape(search_clean_html), ocr_text_clean, re.IGNORECASE)
                    if exact_match_clean:
                        found_match = exact_match_clean
                        match_quality = 3
                        # แปลงตำแหน่งกลับไปหาใน original text
                        # หาโดยนับจำนวนตัวอักษรที่ผ่านมา
                        char_count_before_match = len(ocr_text_clean[:exact_match_clean.start()])
                        # ค้นหาใน original text โดยใช้ char count ที่ประมาณ
                        approx_start = 0
                        clean_idx = 0
                        for orig_idx, char in enumerate(ocr_text):
                            if clean_idx >= char_count_before_match:
                                approx_start = orig_idx
                                break
                            if char not in ['<', '>'] and not (orig_idx > 0 and ocr_text[orig_idx-1:orig_idx+1] == '</'):
                                clean_idx += 1
                        # สร้าง match object ใหม่ที่มีตำแหน่งใน original text
                        match_end = approx_start + len(search_clean_html)
                        # ใช้ original text สำหรับการคำนวณตำแหน่ง
                        from collections import namedtuple
                        MatchObj = namedtuple('MatchObj', ['start', 'end', 'group'])
                        found_match = MatchObj(start=approx_start, end=match_end, group=lambda: search_clean_html)
            
            # วิธีที่ 2: Exact match ใน original text (ถ้ายังไม่เจอ)
            if not found_match or match_quality < 2:
                exact_match_orig = re.search(re.escape(search_for_matching), ocr_text, re.IGNORECASE)
                if exact_match_orig:
                    found_match = exact_match_orig
                    match_quality = 2
            
            # วิธีที่ 3: Partial match ที่ครอบคลุมมากที่สุด (สำหรับข้อความยาวๆ หรือเมื่อ exact ไม่เจอ)
            if not found_match or (match_quality == 0 and len(search_clean_html) > 30):
                # สำหรับข้อความยาวๆ ให้หา pattern ที่ครอบคลุมมากที่สุด
                # ใช้คำสำคัญหลายคำเพื่อหา pattern ที่ถูกต้อง
                words = search_clean_html.split()
                if len(words) >= 3:
                    # ลองหลายแบบ: 3 คำแรก, 5 คำแรก, หรือคำสำคัญ
                    for num_words in [min(10, len(words)), min(7, len(words)), min(5, len(words)), min(3, len(words))]:
                        key_phrase = ' '.join(words[:num_words])
                        pattern = re.escape(key_phrase)
                        partial_match = re.search(pattern, ocr_text_clean, re.IGNORECASE)
                        if partial_match:
                            found_match = partial_match
                            match_quality = 1
                            break
            
            # ถ้ายังไม่เจอ ให้ลองหาใน original text (ไม่ clean)
            if not found_match:
                if search_clean_html:
                    final_match = re.search(re.escape(search_clean_html[:min(50, len(search_clean_html))]), ocr_text, re.IGNORECASE)
                    if final_match:
                        found_match = final_match
                        match_quality = 1
            
            if found_match:
                # สำหรับ scan PDF เราไม่สามารถหา exact position ได้
                # แต่เราสามารถสร้าง approximate positions โดยแบ่ง PDF เป็น grid
                # หรือแสดง highlight แบบ overlay ที่ประมาณตำแหน่ง
                
                # ใช้ PyMuPDF เพื่อหา page dimensions
                if HAS_PYMUPDF:
                    try:
                        doc = fitz.open(pdf_path)
                        if page_num > 0 and page_num <= len(doc):
                            page = doc[page_num - 1]
                            page_rect = page.rect
                            
                            # ใช้ text position ใน OCR text เพื่อประมาณตำแหน่งใน PDF
                            # แบ่ง OCR text เป็น lines และหา line ที่มี text
                            ocr_lines = ocr_text.split('\n')
                            
                            # หา line number ของ match ใน original text
                            match_start = found_match.start()
                            line_start = ocr_text[:match_start].count('\n')
                            match_end = found_match.end()
                            
                            # หา line ที่ครอบคลุม match (อาจมีหลายบรรทัด)
                            end_line = ocr_text[:match_end].count('\n')
                            num_lines = max(1, end_line - line_start + 1)
                            
                            # สำหรับ Sales Promotion - ตรวจสอบว่าต้องไม่ match กับ "หมายเหตุ..."
                            # ถ้า match กับ line ที่มี "หมายเหตุ" ให้ข้ามไปหา line ถัดไป
                            if line_start < len(ocr_lines):
                                current_line = ocr_lines[line_start]
                                if 'หมายเหตุ' in current_line and 'ค่าส่งเสริมการขาย' in search_clean_html:
                                    # ถ้า match กับ line ที่มี "หมายเหตุ" ให้หา line ถัดไปที่มี "ค่าส่งเสริมการขาย"
                                    for next_line_idx in range(line_start + 1, min(line_start + 3, len(ocr_lines))):
                                        if 'ค่าส่งเสริมการขาย' in ocr_lines[next_line_idx]:
                                            line_start = next_line_idx
                                            # ปรับ match_end ให้ครอบคลุมทั้งบรรทัด
                                            match_text = found_match.group(0) if hasattr(found_match, 'group') else search_clean_html
                                            num_lines = 1
                                            break
                            
                            # นับจำนวนบรรทัดที่มีข้อความจริงๆ (non-empty lines)
                            total_lines = len([l for l in ocr_lines if l.strip()])
                            if total_lines > 0:
                                # คำนวณ Y position (จากบนลงล่าง)
                                page_height = page_rect.height
                                page_width = page_rect.width
                                top_margin = 50
                                bottom_margin = 50
                                content_height = page_height - top_margin - bottom_margin
                                
                                # คำนวณ Y position จากบรรทัดแรกของ match
                                line_ratio_start = line_start / max(total_lines, 1)
                                y0 = top_margin + (line_ratio_start * content_height)
                                
                                # คำนวณความสูงของ box ตามจำนวนบรรทัด (ประมาณ 18 points ต่อบรรทัด)
                                box_height = max(18, num_lines * 18)
                                
                                # ตรวจสอบว่ามีหลายบรรทัดหรือไม่
                                if num_lines > 1:
                                    line_ratio_end = min(line_start + num_lines, total_lines) / max(total_lines, 1)
                                    y1 = top_margin + (line_ratio_end * content_height)
                                    box_height = y1 - y0
                                else:
                                    y1 = y0 + box_height
                                
                                # คำนวณ X position จากตำแหน่งในบรรทัดแรก
                                if line_start < len(ocr_lines):
                                    line_text = ocr_lines[line_start]
                                    # ลบ HTML tags จาก line_text เพื่อหา position ที่ถูกต้อง
                                    line_text_clean = re.sub(r'<[^>]+>', '', line_text)
                                    
                                    # หาตำแหน่งของ match ในบรรทัด
                                    match_in_line = found_match.group(0)
                                    # ลองหาใน line_text (อาจมี HTML tags)
                                    line_pos = line_text.find(match_in_line[:min(20, len(match_in_line))])
                                    if line_pos == -1:
                                        # ถ้าไม่เจอ ลองหาใน clean version
                                        match_in_line_clean = re.sub(r'<[^>]+>', '', match_in_line)
                                        line_pos = line_text_clean.find(match_in_line_clean[:min(20, len(match_in_line_clean))])
                                    
                                    if line_pos >= 0:
                                        char_ratio = line_pos / max(len(line_text_clean), 1) if line_text_clean else 0.3
                                    else:
                                        char_ratio = 0.3
                                else:
                                    char_ratio = 0.3
                                
                                left_margin = 50
                                right_margin = 50
                                content_width = page_width - left_margin - right_margin
                                
                                x0 = left_margin + (char_ratio * content_width)
                                
                                # คำนวณความกว้างตามจำนวนตัวอักษร (ประมาณ 6-7 points ต่อตัวอักษรสำหรับ font ขนาดปกติ)
                                # สำหรับข้อความยาวๆ ให้ใช้สัดส่วนของ content_width
                                actual_search_text = search_clean_html if search_clean_html else search_for_matching
                                text_length = len(actual_search_text)
                                
                                # ตรวจสอบว่าเป็น Description หรือไม่ (อยู่ใน table)
                                is_description = 'รายการ' in actual_search_text or 'Promotion' in actual_search_text or line_start < len(ocr_lines) and '<table>' in '\n'.join(ocr_lines[max(0, line_start-5):line_start+1])
                                
                                # สำหรับ Description หรือข้อความยาวๆ (>30 ตัวอักษร)
                                if is_description or text_length > 30:
                                    # ใช้สัดส่วนที่ใหญ่ขึ้นเพื่อครอบคลุมข้อความทั้งหมด
                                    # สำหรับ Description ใน table ให้ใช้ 65-75% ของ content width
                                    if is_description:
                                        # คำนวณความกว้างตามจำนวนตัวอักษร แต่ไม่เกิน 75% ของ content width
                                        estimated_width = text_length * 6.5  # 6.5 points ต่อตัวอักษร
                                        text_width = min(estimated_width, content_width * 0.75)
                                        # จำกัดให้ไม่เกิน 75% และไม่น้อยกว่า 60%
                                        text_width = max(content_width * 0.6, min(text_width, content_width * 0.75))
                                    else:
                                        # สำหรับข้อความยาวๆ ทั่วไป
                                        text_width = min(content_width * 0.7, content_width - (char_ratio * content_width))
                                else:
                                    # สำหรับข้อความสั้นๆ ใช้การคำนวณตามจำนวนตัวอักษร
                                    text_width = min(text_length * 7, content_width * 0.6)
                                
                                x1 = x0 + text_width
                                
                                # จำกัดให้อยู่ในขอบเขต
                                x0 = max(left_margin, min(x0, page_width - right_margin - 50))
                                x1 = max(x0 + 50, min(x1, page_width - right_margin))
                                y0 = max(top_margin, min(y0, page_height - bottom_margin - box_height))
                                y1 = max(y0 + 18, min(y1, page_height - bottom_margin))
                                box_height = y1 - y0
                                
                                pos = {
                                    'x0': float(x0),
                                    'y0': float(y0),
                                    'x1': float(x1),
                                    'y1': float(y1),
                                    'page': page_num,
                                    'text': search_for_matching,
                                    'method': 'ocr_txt_approximate',
                                    'page_width': float(page_width),
                                    'page_height': float(page_height)
                                }
                                
                                # สำหรับ Description field: ขยาย bounding box
                                if field_name and 'description' in str(field_name).lower():
                                    new_x0 = 40.0  # x0 = 40 (คงที่)
                                    # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                                    original_width_ratio = 0.60  # 70% - 10% = 60%
                                    new_x1 = new_x0 + (page_width * original_width_ratio)
                                    line_height = 20
                                    # เลื่อนขึ้นไปขอบกระดาษบน (y0 = 0 หรือใกล้ 0)
                                    new_y0 = 290.0  # y0 = 290
                                    new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                                    new_x0 = max(0, min(new_x0, page_width - 100))
                                    new_x1 = max(new_x0 + 100, min(new_x1, page_width))
                                    new_y1 = max(new_y0 + 20, min(new_y1, page_height))
                                    pos['x0'] = float(new_x0)
                                    pos['x1'] = float(new_x1)
                                    pos['y0'] = float(new_y0)
                                    pos['y1'] = float(new_y1)
                                
                                positions.append(pos)
                            else:
                                # Fallback: ใช้ตำแหน่งกลาง
                                center_x = page_rect.width / 2
                                center_y = page_rect.height / 2
                                # สำหรับข้อความยาวๆ ให้ใช้ความกว้างที่เหมาะสม
                                actual_search_text = search_clean_html if search_clean_html else search_for_matching
                                highlight_width = min(page_rect.width * 0.6, len(actual_search_text) * 7)
                                highlight_height = max(30, min(100, (actual_search_text.count(' ') + 1) * 18))
                                
                                pos = {
                                    'x0': float(center_x - highlight_width / 2),
                                    'y0': float(center_y - highlight_height / 2),
                                    'x1': float(center_x + highlight_width / 2),
                                    'y1': float(center_y + highlight_height / 2),
                                    'page': page_num,
                                    'text': search_for_matching,
                                    'method': 'ocr_txt_center',
                                    'page_width': float(page_rect.width),
                                    'page_height': float(page_rect.height)
                                }
                                
                                # สำหรับ Description field: ขยาย bounding box
                                if field_name and 'description' in str(field_name).lower():
                                    page_width = pos['page_width']
                                    page_height = pos['page_height']
                                    new_x0 = 40.0  # x0 = 40 (คงที่)
                                    # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
                                    original_width_ratio = 0.60  # 70% - 10% = 60%
                                    new_x1 = new_x0 + (page_width * original_width_ratio)
                                    line_height = 20
                                    # เลื่อนขึ้นไปขอบกระดาษบน (y0 = 0 หรือใกล้ 0)
                                    new_y0 = 290.0  # y0 = 290
                                    new_y1 = new_y0 + (line_height * 5.5)  # สูงขึ้นเล็กน้อย
                                    new_x0 = max(0, min(new_x0, page_width - 100))
                                    new_x1 = max(new_x0 + 100, min(new_x1, page_width))
                                    new_y1 = max(new_y0 + 20, min(new_y1, page_height))
                                    pos['x0'] = float(new_x0)
                                    pos['x1'] = float(new_x1)
                                    pos['y0'] = float(new_y0)
                                    pos['y1'] = float(new_y1)
                                
                                positions.append(pos)
                            
                        doc.close()
                    except Exception as e:
                        st.warning(f"⚠️ Could not get PDF dimensions: {e}")
                else:
                    # ถ้าไม่มี PyMuPDF ให้ใช้ default approximate position
                    positions.append({
                        'x0': 100.0,
                        'y0': 100.0,
                        'x1': 400.0,
                        'y1': 150.0,
                        'page': page_num,
                        'text': clean_search,
                        'method': 'ocr_txt_default'
                    })
        else:
            # ไม่พบไฟล์ .txt
            st.warning(f"⚠️ OCR text file not found: {txt_filename}. Please run OCR first.")
    
    except Exception as e:
        st.warning(f"⚠️ Error reading OCR text file: {e}")
    
    # Validate และ filter bounding boxes ที่ผิดปกติ
    validated_positions = []
    for pos in positions:
        width = pos['x1'] - pos['x0']
        height = pos['y1'] - pos['y0']
        
        # Get page dimensions for validation
        page_width = pos.get('page_width', 595)  # A4 default width in points
        page_height = pos.get('page_height', 842)  # A4 default height in points
        
        # สำหรับ Description field: ขยาย bounding box ให้ครอบคลุมข้อความทั้งหมด
        # - ความกว้าง: x0 = 40 (คงที่), x1 ปรับตามสัดส่วน
        # - ความสูง: เลื่อนขึ้นไปขอบกระดาษบน
        if field_name and 'description' in str(field_name).lower():
            # x0 = 40 (คงที่)
            new_x0 = 40.0
            # คำนวณ x1 ตามสัดส่วนเดิม (จาก 0.10 ถึง 0.70 = 60% ของความกว้าง)
            original_width_ratio = 0.60  # 70% - 10% = 60%
            new_x1 = new_x0 + (page_width * original_width_ratio)
            
            # ขยายความสูง: ครอบคลุม 5.5 บรรทัด (ประมาณ 110 points)
            # y0 = 290 (เลื่อนลงมา)
            line_height = 20  # ความสูงต่อบรรทัด (ประมาณ)
            new_y0 = 290.0  # y0 = 290
            new_y1 = new_y0 + (line_height * 5.5)  # 5.5 บรรทัด
            
            # จำกัดให้อยู่ในขอบเขตของหน้า
            new_x0 = max(0, min(new_x0, page_width - 100))
            new_x1 = max(new_x0 + 100, min(new_x1, page_width))
            new_y1 = max(new_y0 + 20, min(new_y1, page_height))
            
            # อัพเดท bounding box
            pos['x0'] = float(new_x0)
            pos['x1'] = float(new_x1)
            pos['y0'] = float(new_y0)
            pos['y1'] = float(new_y1)
            
            # Recalculate width and height
            width = pos['x1'] - pos['x0']
            height = pos['y1'] - pos['y0']
        
        # Filter out boxes ที่ผิดปกติ:
        # - ต้องมี width และ height เป็นบวก
        # - width ต้องไม่เกิน 90% ของหน้า (รองรับข้อความยาวๆ)
        # - height ต้องไม่เกิน 50% ของหน้า (รองรับข้อความหลายบรรทัด)
        # - ต้องอยู่ในขอบเขตของหน้า
        max_width = page_width * 0.9
        max_height = page_height * 0.5
        
        if (width > 0 and height > 0 and 
            width <= max_width and height <= max_height and
            pos['x0'] >= 0 and pos['y0'] >= 0 and
            pos['x1'] <= page_width and pos['y1'] <= page_height):
            validated_positions.append(pos)
        else:
            # Debug: แสดงเฉพาะใน console ไม่แสดงใน UI
            print(f"[DEBUG] Filtered invalid bbox: width={width:.1f}, height={height:.1f}, "
                  f"page_size=({page_width:.0f}x{page_height:.0f}), method={pos.get('method', 'unknown')}")
    
    return validated_positions

def render_pdf(file_path, page_num=1, highlight_positions=None, zoom_level=1.0):
    """
    Render PDF as image with highlight boxes using pdf2image + PIL
    highlight_positions: list of dicts with 'x0', 'y0', 'x1', 'y1', 'page', 'page_width', 'page_height'
    zoom_level: float, 1.0 = 100%, 1.5 = 150%, 0.5 = 50%, etc.
    """
    if not os.path.exists(file_path):
        st.error(f"PDF file not found: {file_path}")
        return
    
    try:
        from pdf2image import convert_from_path
        from PIL import Image, ImageDraw
        
        # Convert PDF page to image
        # ใช้ DPI สูงขึ้นเมื่อ zoom in เพื่อความชัดเจน
        base_dpi = 150
        effective_dpi = int(base_dpi * max(1.0, zoom_level))
        images = convert_from_path(
            file_path,
            first_page=page_num,
            last_page=page_num,
            dpi=effective_dpi,
            poppler_path=POPPLER_PATH if POPPLER_PATH else None
        )
        
        if not images:
            st.error(f"Failed to render PDF page {page_num}")
            return
        
        img = images[0]
        
        # Apply zoom if needed
        if zoom_level != 1.0:
            original_size = img.size
            new_size = (int(original_size[0] * zoom_level), int(original_size[1] * zoom_level))
            # ใช้ LANCZOS resampling สำหรับคุณภาพที่ดี
            img = img.resize(new_size, Image.Resampling.LANCZOS)
        
        # Draw highlight boxes if provided
        if highlight_positions and len(highlight_positions) > 0:
            draw = ImageDraw.Draw(img)
            img_width, img_height = img.size
            
            # Get page dimensions from first position
            first_pos = highlight_positions[0]
            page_width = first_pos.get('page_width', 595)
            page_height = first_pos.get('page_height', 842)
            
            # Scale factors from PDF coordinates to image coordinates
            scale_x = img_width / page_width
            scale_y = img_height / page_height
            
            boxes_drawn = 0
            for pos in highlight_positions:
                if pos.get('page') != page_num:
                    continue
                
                # Convert PDF coordinates to image pixel coordinates
                x0 = pos['x0'] * scale_x
                y0 = pos['y0'] * scale_y
                x1 = pos['x1'] * scale_x
                y1 = pos['y1'] * scale_y
                
                # Validate coordinates (must have positive width and height)
                if x1 <= x0 or y1 <= y0:
                    print(f"[DEBUG] Skipping invalid box: x0={x0:.1f}, y0={y0:.1f}, x1={x1:.1f}, y1={y1:.1f}")
                    continue
                
                # เพิ่ม padding รอบๆ กรอบ (8 pixels แต่ละด้าน)
                padding = 8
                x0 = x0 - padding
                y0 = y0 - padding
                x1 = x1 + padding
                y1 = y1 + padding
                
                # Ensure boxes are within image bounds
                x0 = max(0, min(x0, img_width - 1))
                y0 = max(0, min(y0, img_height - 1))
                x1 = max(x0 + 10, min(x1, img_width))
                y1 = max(y0 + 10, min(y1, img_height))
                
                # Double-check after clamping
                if x1 <= x0 or y1 <= y0:
                    continue
                
                try:
                    # Draw thick red rectangle (multiple passes for thickness)
                    for i in range(5):
                        draw.rectangle(
                            [(x0-i, y0-i), (x1+i, y1+i)],
                            outline='red',
                            width=1
                        )
                    # Inner semi-transparent effect
                    if x1 > x0 + 6 and y1 > y0 + 6:
                        draw.rectangle(
                            [(x0+3, y0+3), (x1-3, y1-3)],
                            outline='#FF8888',
                            width=2
                        )
                    boxes_drawn += 1
                except Exception as e:
                    print(f"[DEBUG] Error drawing box: {e}")
            
        # Display the rendered image
        # ใช้ HTML/CSS เพื่อบังคับขนาดจริงตาม zoom level
        img_width, img_height = img.size
        
        if zoom_level != 1.0:
            # แปลง PIL Image เป็น base64 เพื่อใช้ใน HTML
            import io
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
            
            # ใช้ HTML/CSS เพื่อบังคับขนาดจริง (ไม่ให้ Streamlit resize)
            zoom_html = f"""
            <div style="text-align: center; margin: 10px 0;">
                <img src="data:image/png;base64,{img_base64}" 
                     style="max-width: none; width: {img_width}px; height: auto; display: block; margin: 0 auto;"
                     alt="PDF Page {page_num}">
                <p style="margin-top: 5px; color: #888;">📄 Page {page_num} (Zoom: {int(zoom_level * 100)}%)</p>
            </div>
            """
            st.markdown(zoom_html, unsafe_allow_html=True)
        else:
            # ที่ 100% ใช้ use_container_width เพื่อให้พอดี container
            st.image(img, use_container_width=True, caption=f"📄 Page {page_num} (Zoom: {int(zoom_level * 100)}%)")
        
    except ImportError as e:
        st.error(f"❌ Missing required library: {e}")
        st.info("📦 Install with: `pip install pdf2image Pillow`")
    except Exception as e:
        st.error(f"Error rendering PDF: {e}")
        # Fallback to browser iframe mode (no highlight support)
        try:
            with open(file_path, "rb") as f: 
                pdf_data = f.read()
            if pdf_data:
                base64_pdf = base64.b64encode(pdf_data).decode('utf-8')
                pdf_html = f'<iframe src="data:application/pdf;base64,{base64_pdf}#page={page_num}&toolbar=1" width="100%" height="800px" style="border: none;"></iframe>'
                st.markdown(f"**Browser Mode (Page {page_num}) - No highlight support**")
                st.markdown(pdf_html, unsafe_allow_html=True)
        except:
            st.error("Failed to render PDF in any mode")

def extract_hyperlinks(source, sheet_name):
    """
    Extract hyperlinks from Excel, returning both target path and display text
    Returns: list of dicts with structure {column_name: {'target': path, 'display': text}}
    """
    try:
        if hasattr(source, 'seek'): 
            source.seek(0)
        wb = openpyxl.load_workbook(source, data_only=False)
        ws = wb[sheet_name]
        link_data = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            row_links = {}
            for cell in row:
                header_val = ws.cell(row=1, column=cell.column).value
                target = None
                display = None
                
                if cell.hyperlink: 
                    target = cell.hyperlink.target
                    display = cell.value if cell.value else target
                elif isinstance(cell.value, str) and str(cell.value).strip().upper().startswith("=HYPERLINK"):
                    # Parse HYPERLINK formula: =HYPERLINK("target", "display")
                    formula = cell.value.strip()
                    # Extract both target and display text
                    matches = re.findall(r'["\']([^"\']+)["\']', formula)
                    if len(matches) >= 1:
                        target = matches[0]
                        display = matches[1] if len(matches) >= 2 else target
                    
                if target: 
                    row_links[header_val] = {
                        'target': target,
                        'display': display if display else target
                    }
            link_data.append(row_links)
        wb.close()
        return link_data
    except Exception as e:
        print(f"[DEBUG] Error extracting hyperlinks: {e}")
        return []

def get_sheet_names_fresh(file_source):
    try:
        if hasattr(file_source, 'seek'): 
            file_source.seek(0)
        wb = openpyxl.load_workbook(file_source, read_only=True, keep_links=False)
        return wb.sheetnames
    except: 
        return []

# --- LOOKUP LOGIC ---
def load_vendor_master(force_reload=False):
    """
    โหลด Vendor Master จาก Vendor_branch.xlsx
    force_reload: ถ้าเป็น True จะ reload ใหม่แม้จะมีข้อมูลใน session state แล้ว
    """
    # ตรวจสอบว่าไฟล์มีอยู่หรือไม่
    if not os.path.exists(VENDOR_MASTER_PATH):
        st.error(f"❌ ไม่พบไฟล์ Vendor Master ที่: {VENDOR_MASTER_PATH}")
        st.session_state.vendor_master_df = None
        return None
    
    # ตรวจสอบว่าไฟล์ถูกแก้ไขหรือไม่ (โดยใช้ mtime)
    file_mtime = os.path.getmtime(VENDOR_MASTER_PATH)
    
    # เก็บ mtime ไว้ใน session state เพื่อตรวจสอบการเปลี่ยนแปลง
    if 'vendor_master_mtime' not in st.session_state:
        st.session_state.vendor_master_mtime = 0
    
    # ถ้าไฟล์ถูกแก้ไข หรือ force_reload หรือยังไม่มีข้อมูลใน session state
    need_reload = (
        force_reload or 
        st.session_state.vendor_master_mtime != file_mtime or
        'vendor_master_df' not in st.session_state or 
        st.session_state.vendor_master_df is None
    )
    
    if need_reload:
        try:
            # ตรวจสอบ sheets ทั้งหมดและเลือก sheet ที่มีข้อมูลมากที่สุด
            xl_file = pd.ExcelFile(VENDOR_MASTER_PATH)
            
            # ถ้ามีหลาย sheets ให้เลือก sheet แรก (หรือ sheet ที่มีข้อมูลมากที่สุด)
            max_rows = 0
            best_sheet = xl_file.sheet_names[0]
            
            for sheet_name in xl_file.sheet_names:
                temp_df = pd.read_excel(VENDOR_MASTER_PATH, sheet_name=sheet_name, nrows=1)
                if not temp_df.empty:
                    full_df = pd.read_excel(VENDOR_MASTER_PATH, sheet_name=sheet_name)
                    if len(full_df) > max_rows:
                        max_rows = len(full_df)
                        best_sheet = sheet_name
            
            df = pd.read_excel(VENDOR_MASTER_PATH, sheet_name=best_sheet, dtype=str)
            df.columns = df.columns.str.strip()  # Clean column names
            
            # Clean Data: เลขผู้เสียภาษี (ลบขีด ลบวรรค) - เหมือน Extract_Inv.py
            if 'เลขประจำตัวผู้เสียภาษี' in df.columns: 
                df['เลขประจำตัวผู้เสียภาษี'] = df['เลขประจำตัวผู้เสียภาษี'].fillna('').str.replace(r'\D', '', regex=True)
            
            # Clean Data: สาขา (เติม 0 ให้ครบ 5 หลัก และแปลง "สำนักงานใหญ่" เป็น "00000") - เหมือน Extract_Inv.py
            if 'สาขา' in df.columns: 
                def clean_branch(x):
                    x = str(x).strip()
                    # แปลง "สำนักงานใหญ่" หรือ "Head Office" เป็น "00000"
                    if re.search(r'(สำนักงานใหญ่|สนญ|Head\\s*Office|H\\.\\?O\\.\\?)', x, re.IGNORECASE):
                        return "00000"
                    # ถ้าเป็นตัวเลข ให้เติม 0 ให้ครบ 5 หลัก
                    if x.isdigit():
                        return x.zfill(5)
                    return x
                df['สาขา'] = df['สาขา'].fillna('').apply(clean_branch)
            
            st.session_state.vendor_master_df = df
            st.session_state.vendor_master_mtime = file_mtime
            st.toast(f"✅ โหลด Vendor Master สำเร็จ ({len(df)} rows) จาก sheet: {best_sheet}", icon="✅")
        except Exception as e:
            st.toast(f"⚠️ อ่านไฟล์ Master Error: {e}", icon="⚠️")
            st.session_state.vendor_master_df = None
            return None
    
    return st.session_state.vendor_master_df

def lookup_vendor_info(vendor_id, branch, debug=False):
    """
    Lookup Vendor code จาก Vendor Master โดยใช้:
    - vendor_id (VendorID_OCR) = เลขประจำตัวผู้เสียภาษี
    - branch (Branch_OCR) = สาขา
    Return: {'code': 'Vendor code SAP', 'name': 'ชื่อบริษัท'} หรือ None
    """
    # Reload vendor master เพื่อให้ได้ข้อมูลล่าสุด (force_reload=False จะ reload เฉพาะเมื่อไฟล์ถูกแก้ไข)
    master = load_vendor_master(force_reload=False)
    if master is None:
        if debug:
            st.error(f"DEBUG: ไม่สามารถโหลด Vendor Master ได้ - Path: {VENDOR_MASTER_PATH}")
            if not os.path.exists(VENDOR_MASTER_PATH):
                st.error(f"DEBUG: ไฟล์ไม่พบที่ path: {VENDOR_MASTER_PATH}")
        return None
    
    try:
        # Clean input ให้เหมือนกับที่ clean ใน Master
        # Clean vendor_id: ลบขีด ลบวรรค (เหมือนใน Extract_Inv.py)
        v_id_original = vendor_id
        v_id = str(vendor_id).strip() if vendor_id and pd.notna(vendor_id) else ""
        v_id = re.sub(r'\D', '', v_id)  # ลบทุกอย่างที่ไม่ใช่ตัวเลข
        
        # Clean branch: เติม 0 ให้ครบ 5 หลัก และแปลง "สำนักงานใหญ่" เป็น "00000" (เหมือนใน Extract_Inv.py)
        def clean_branch_input(x):
            if x is None or pd.isna(x):
                return ""
            x = str(x).strip()
            # แปลง "สำนักงานใหญ่" หรือ "Head Office" เป็น "00000"
            if re.search(r'(สำนักงานใหญ่|สนญ|Head\s*Office|H\.?O\.?)', x, re.IGNORECASE):
                return "00000"
            # ถ้าเป็นตัวเลข ให้เติม 0 ให้ครบ 5 หลัก
            if x.isdigit():
                return x.zfill(5)
            return x
        
        br_original = branch
        br = clean_branch_input(branch)
        
        if debug:
            st.info(f"DEBUG: Input - vendor_id='{v_id_original}' → cleaned='{v_id}', branch='{br_original}' → cleaned='{br}'")
        
        # ตรวจสอบว่ามีค่าที่จะ search หรือไม่
        if not v_id or not br:
            if debug:
                st.warning(f"DEBUG: ข้อมูลไม่ครบ - v_id={v_id}, br={br}")
            return None
        
        # ตรวจสอบว่า Master มี column ที่ต้องการหรือไม่
        if 'เลขประจำตัวผู้เสียภาษี' not in master.columns or 'สาขา' not in master.columns:
            if debug:
                st.error(f"DEBUG: ขาด column ใน Master - columns={list(master.columns)}")
            return None
        
        # หา column name ของ Vendor code (อาจเป็น "Vendor code SAP" หรือ "Vendor code SA")
        vendor_code_col = None
        for col in master.columns:
            col_lower = str(col).lower()
            if 'vendor' in col_lower and 'code' in col_lower:
                vendor_code_col = col
                break
        
        if not vendor_code_col:
            if debug:
                st.error(f"DEBUG: ไม่พบ Vendor code column ใน Master - columns={list(master.columns)}")
            return None
        
        if debug:
            st.info(f"DEBUG: Using vendor_code_col='{vendor_code_col}'")
            st.info(f"DEBUG: Master has {len(master)} total rows")
            st.info(f"DEBUG: Searching for v_id='{v_id}' AND br='{br}' in Master")
            
            # แสดงตัวอย่างข้อมูลใน Master
            st.markdown("**📊 ตัวอย่างข้อมูลใน Master (5 rows แรก):**")
            sample = master[['เลขประจำตัวผู้เสียภาษี', 'สาขา', vendor_code_col]].head(5)
            st.dataframe(sample, use_container_width=True)
            
            # ค้นหาว่ามี vendor_id นี้ใน Master หรือไม่ (ไม่สนใจ branch)
            vendor_matches = master[master['เลขประจำตัวผู้เสียภาษี'] == v_id]
            if not vendor_matches.empty:
                st.warning(f"DEBUG: พบ vendor_id='{v_id}' ใน Master ({len(vendor_matches)} rows) แต่ branch ไม่ตรง:")
                vendor_sample = vendor_matches[['เลขประจำตัวผู้เสียภาษี', 'สาขา', vendor_code_col]].head(10)
                st.dataframe(vendor_sample, use_container_width=True)
                st.info(f"💡 คุณกำลังหา branch='{br}' แต่ใน Master มี branch: {list(vendor_matches['สาขา'].unique())}")
            else:
                st.error(f"DEBUG: ไม่พบ vendor_id='{v_id}' ใน Master เลย")
                # แสดง vendor_id ที่ใกล้เคียง (ขึ้นต้นเหมือนกัน)
                prefix = v_id[:8] if len(v_id) >= 8 else v_id[:4]
                similar = master[master['เลขประจำตัวผู้เสียภาษี'].str.startswith(prefix)]
                if not similar.empty:
                    st.info(f"💡 พบ vendor_id ที่ขึ้นต้นด้วย '{prefix}' ({len(similar)} rows):")
                    similar_sample = similar[['เลขประจำตัวผู้เสียภาษี', 'สาขา', vendor_code_col]].head(10)
                    st.dataframe(similar_sample, use_container_width=True)
        
        # Search ใน Master - ใช้ exact match
        res = master[(master['เลขประจำตัวผู้เสียภาษี'] == v_id) & (master['สาขา'] == br)]
        
        if debug:
            st.info(f"DEBUG: Found {len(res)} matching records")
            if not res.empty:
                st.dataframe(res, use_container_width=True)
        
        result = {}
        if not res.empty:
            # ดึง Vendor code SAP
            vendor_code = str(res.iloc[0][vendor_code_col]).strip()
            # ตรวจสอบว่าไม่ใช่ nan หรือ empty
            if vendor_code and vendor_code.lower() != 'nan' and vendor_code != '':
                result['code'] = vendor_code
                if debug:
                    st.success(f"DEBUG: Found Vendor code = '{vendor_code}'")
            
            # ดึงชื่อบริษัท (ถ้ามี)
            if 'ชื่อบริษัท' in res.columns: 
                vendor_name = str(res.iloc[0]['ชื่อบริษัท']).strip()
                if vendor_name and vendor_name.lower() != 'nan' and vendor_name != '':
                    result['name'] = vendor_name
        else:
            if debug:
                st.warning("DEBUG: ไม่พบข้อมูลที่ตรงกัน")
        
        return result if result else None
    except Exception as e:
        if debug:
            st.error(f"DEBUG: Error in lookup_vendor_info: {e}")
        import traceback
        if debug:
            st.code(traceback.format_exc())
    return None

def find_column_name(columns, keywords):
    cols_lower = [str(c).lower() for c in columns]
    for col, col_lower in zip(columns, cols_lower):
        if all(k.lower() in col_lower for k in keywords): 
            return col
    return None

def format_date_value(value, column_name):
    """
    Format date value to dd/MM/yyyy format for InvDateOCR columns.
    Removes time portion if present.
    """
    if "InvDateOCR" in str(column_name):
        value_str = str(value).strip()
        if not value_str or value_str.lower() in ['nan', 'none', '']:
            return value_str
        
        # Try to parse datetime and format to dd/MM/yyyy
        try:
            # Check if it contains time portion (e.g., "2025-09-13 00:00:00" or "2025-08-25 00:00:00")
            if ' ' in value_str and len(value_str) > 10:
                # Extract date part before space
                date_part = value_str.split(' ')[0]
                # Try to parse as yyyy-MM-dd and convert to dd/MM/yyyy
                try:
                    dt = datetime.strptime(date_part, '%Y-%m-%d')
                    return dt.strftime('%d/%m/%Y')
                except ValueError:
                    # If not yyyy-MM-dd, try pandas parsing
                    dt = pd.to_datetime(date_part, errors='coerce')
                    if pd.notna(dt):
                        return dt.strftime('%d/%m/%Y')
            # If already in yyyy-MM-dd format, convert to dd/MM/yyyy
            elif re.match(r'^\d{4}-\d{2}-\d{2}$', value_str):
                dt = datetime.strptime(value_str, '%Y-%m-%d')
                return dt.strftime('%d/%m/%Y')
            # If already in dd/MM/yyyy format, return as is
            elif re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', value_str):
                # Validate the date format
                try:
                    dt = pd.to_datetime(value_str, format='%d/%m/%Y', errors='coerce')
                    if pd.notna(dt):
                        return value_str
                except:
                    pass
            # Try to parse as datetime with pandas (handles various formats)
            dt = pd.to_datetime(value_str, errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%d/%m/%Y')
        except (ValueError, AttributeError, TypeError):
            pass
    
    return value

def format_number_value(value, column_name):
    """
    Format number value to 999,999.99 format for InvAmtOCR columns.
    Adds thousand separators (commas) and ensures 2 decimal places.
    """
    if "InvAmtOCR" in str(column_name):
        value_str = str(value).strip()
        if not value_str or value_str.lower() in ['nan', 'none', '']:
            return value_str
        
        try:
            # Remove any existing commas and spaces
            cleaned_value = value_str.replace(',', '').replace(' ', '')
            
            # Try to convert to float
            num_value = float(cleaned_value)
            
            # Format with 2 decimal places and thousand separators
            formatted = f"{num_value:,.2f}"
            return formatted
        except (ValueError, TypeError):
            # If conversion fails, return original value
            return value_str
    
    return value

def generate_sap_data(df_source):
    try:
        # Rules (Shortened)
        column_rules = [
            ["Running", ""], ["Fix", "1001"], ["Col", "InvDateOCR"], ["Fix", "KR"],
            ["Fix", "Now Date"], ["Fix", "THB"], ["Col", "InvNoOCR"], ["Fix", "ค่าผ่านท่า"],
            ["Fix", "0000"], ["Fix", "31"], ["Col", "Vendor Match"], ["Col", "InvAmtOCR"],
            ["Fix", "V7"], ["Fix", "X"], ["Fix", "0110"], ["Col", "CyOrg"],
            ["Col", "CyInvoiceNo"], ["Fix", "ค่าผ่านท่า"], ["Fix", "6450200"],
        ]
        new_data = {}
        row_count = len(df_source)
        for idx, (rule_type, rule_value) in enumerate(column_rules):
            col_name = f"Col_{idx+1}"
            if rule_type == "Fix":
                val = [datetime.now().strftime("%d%m%Y")] * row_count if rule_value == "Now Date" else [rule_value] * row_count
                new_data[col_name] = val
            elif rule_type == "Col":
                matched_col = rule_value
                if rule_value not in df_source.columns:
                    if rule_value == "Vendor Match":
                        matched_col = find_column_name(df_source.columns, ["vendor", "match"]) or rule_value
                    elif rule_value == "Vendor code":
                        matched_col = find_column_name(df_source.columns, ["vendor", "code"]) or rule_value
                
                if matched_col in df_source.columns:
                    if rule_value == "InvDateOCR":
                        dates = pd.to_datetime(df_source[matched_col], errors='coerce')
                        new_data[col_name] = dates.dt.strftime('%d%m%Y').fillna('')
                    else: 
                        new_data[col_name] = df_source[matched_col].tolist()
                else: 
                    new_data[col_name] = [""] * row_count
            elif rule_type == "Running": 
                new_data[col_name] = range(1, row_count + 1)
        
        df_sap = pd.DataFrame(new_data)
        # ส่งกลับเป็น DataFrame เพื่อไป Save ต่อ
        return True, df_sap
    except Exception as e: 
        return False, str(e)

# --- PAGE 1: AI OCR Dashboard ---
def render_page_1():
    # Top row: "Select Folder for AI OCR" | "เลือกหน้าที่จะทำ AI OCR" | "Page:" | "Settings" - all in one row
    col_title, col_mode, col_page, col_settings = st.columns([0.40, 0.25, 0.20, 0.15])
    with col_title:
        st.title("Select Folder for AI OCR")
    with col_mode:
        # Determine current mode from page_config
        current_config = st.session_state.ocr_page_config
        if current_config == "All":
            default_index = 0
        elif current_config and "-N" in current_config and current_config != "1-N":
            default_index = 1
        elif current_config and "-" in current_config and "-N" not in current_config:
            default_index = 2
        else:
            # If old config is "1-N", convert to "X-N" with X=1
            if current_config == "1-N":
                st.session_state.ocr_page_start = 1
                default_index = 1
            else:
                default_index = 0
        
        page_mode = st.selectbox(
            "เลือกหน้าที่จะทำ AI OCR",
            options=["All", "X-N", "ระบุช่วง"],
            index=default_index,
            key="page_mode_selector",
            label_visibility="visible",
            help="All: เลือกทุกหน้า | X-N: ระบุเริ่มจากหน้า ถึง หน้าสุดท้าย | ระบุช่วง: เลือกตั้งแต่หน้าที่ - ถึงหน้าที่"
        )
        
        # Caption below selectbox
        if page_mode == "All":
            st.caption("📌 เลือกทุกหน้า")
        elif page_mode == "X-N":
            st.caption("📌 ระบุเริ่มจากหน้า X ถึง หน้าสุดท้าย")
        else:  # ระบุช่วง
            st.caption("📌 เลือกตั้งแต่หน้า X ถึง หน้า Y")
    with col_page:
        page_options = {
            "Page 1: AI OCR Dashboard": "🏠",
            "Page 2: Document Editor": "📄"
        }
        selected_page = st.selectbox(
            "Page:",
            options=list(page_options.keys()),
            index=0 if st.session_state.current_page == "Page 1: AI OCR Dashboard" else 1,
            format_func=lambda x: f"{page_options[x]} {x.split(':')[1].strip()}",
            label_visibility="visible"
        )
        if selected_page != st.session_state.current_page:
            st.session_state.current_page = selected_page
            st.rerun()
    
    with col_settings:
        st.markdown("<div style='padding-top: 1.5rem;'>", unsafe_allow_html=True)
        if st.button("⚙️ Settings", use_container_width=True, help="Menu Setting"):
            st.session_state.show_settings = not st.session_state.show_settings
            st.rerun()
        
        # แสดงข้อความตัวเลือก OCR ใต้ปุ่ม Settings
        if st.session_state.ocr_type == "API Typhoon":
            st.caption("🔵 API OCR")
        else:
            st.caption("🟢 Local OCR")
        st.markdown("</div>", unsafe_allow_html=True)
    
    # Settings Dialog
    if st.session_state.show_settings:
        st.markdown("---")
        st.subheader("⚙️ Menu Setting")
        
        col_setting_left, col_setting_right = st.columns([0.5, 0.5])
        
        with col_setting_left:
            st.markdown("### AI OCR Selection")
            ocr_type = st.radio(
                "เลือก AI OCR:",
                options=["API Typhoon", "Local Typhoon"],
                index=0 if st.session_state.ocr_type == "API Typhoon" else 1,
                key="ocr_type_selector",
                help="เลือกประเภท AI OCR ที่ต้องการใช้งาน"
            )
            
            if ocr_type != st.session_state.ocr_type:
                st.session_state.ocr_type = ocr_type
                st.rerun()
            
            st.markdown("---")
            
            # API Configuration - แสดงเฉพาะเมื่อเลือก API Typhoon
            api_key_input = None
            if st.session_state.ocr_type == "API Typhoon":
                st.markdown("### API Configuration")
                api_key_input = st.text_input(
                    "API_KEY:",
                    value=st.session_state.api_key,
                    type="password",
                    help="API Key สำหรับใช้กับ Extract_Inv.py",
                    key="api_key_input"
                )
            # Poppler Path (ใช้ได้ทั้ง API และ Local)
            poppler_path_input = st.text_input(
                "Poppler Path (optional):",
                value=st.session_state.poppler_path or "",
                help="ตัวอย่าง: C:\\poppler\\Library\\bin (ปล่อยว่างเพื่อใช้ค่าจากระบบ)",
                key="poppler_path_input"
            )
            # ไม่แสดง Local Typhoon Configuration ทางซ้าย เพราะมี Information ทางขวาแล้ว
            
            col_save, col_cancel = st.columns([0.5, 0.5])
            with col_save:
                if st.button("💾 Save", use_container_width=True, type="primary"):
                    global POPPLER_PATH
                    poppler_val = poppler_path_input.strip() if poppler_path_input and poppler_path_input.strip() else None
                    if st.session_state.ocr_type == "API Typhoon":
                        if api_key_input and api_key_input.strip():
                            if save_config(api_key_input.strip(), poppler_val):
                                st.session_state.api_key = api_key_input.strip()
                                st.session_state.poppler_path = poppler_val
                                POPPLER_PATH = poppler_val
                                st.success("✅ Settings saved successfully!")
                                time.sleep(1)
                                st.session_state.show_settings = False
                                st.rerun()
                            else:
                                st.error("❌ Failed to save API_KEY")
                        else:
                            st.warning("⚠️ Please enter API_KEY")
                    else:
                        # Local Typhoon - ไม่ต้องบันทึกอะไร แค่ปิด Settings
                        # บันทึกเฉพาะ Poppler Path (เก็บ API_KEY เดิมไว้)
                        if save_config(st.session_state.api_key, poppler_val):
                            st.session_state.poppler_path = poppler_val
                            POPPLER_PATH = poppler_val
                            st.success("✅ Settings saved!")
                            time.sleep(1)
                            st.session_state.show_settings = False
                            st.rerun()
                        else:
                            st.error("❌ Failed to save settings")
            
            with col_cancel:
                if st.button("❌ Cancel", use_container_width=True, type="secondary"):
                    st.session_state.show_settings = False
                    st.rerun()
        
        with col_setting_right:
            st.markdown("### Information")
            if st.session_state.ocr_type == "API Typhoon":
                st.info("""
                    **API Typhoon** ใช้ API จาก OpenTyphoon
                
                - API_KEY จะถูกบันทึกในไฟล์ `config.json`
                - Extract_Inv.py จะอ่าน API_KEY จากไฟล์ config.json อัตโนมัติ
                - ต้องมี API_KEY เพื่อใช้งาน OCR
                - สามารถหา API_KEY ได้ที่: [https://playground.opentyphoon.ai/ocr](https://playground.opentyphoon.ai/ocr)
            """)
            else:
                st.info("""
                **Local Typhoon** ใช้ Ollama ในเครื่อง
                
                - ไม่ต้องใช้ API_KEY
                - ต้องติดตั้ง Ollama และ Model `scb10x/typhoon-ocr1.5-3b:latest`
                - ต้องติดตั้ง Poppler สำหรับแปลง PDF เป็น Image
                - รันผ่าน `Extract_Inv_local.py`
                - ทำงานได้ออฟไลน์ ไม่ต้องเชื่อมต่ออินเทอร์เน็ต
            """)
    
    # Page selection UI based on selected mode - on right side (half screen) below selectbox
    col_left_spacer, col_right_ui = st.columns([0.5, 0.5])
    
    with col_left_spacer:
        pass  # Empty space on left
    
    with col_right_ui:
        # Inputs only (caption moved to above, below selectbox)
        if page_mode == "All":
            page_config = "All"
        elif page_mode == "X-N":
            col_x1, col_x2 = st.columns([0.6, 0.4])
            with col_x1:
                start_page = st.number_input(
                    "หน้าเริ่มต้น:",
                    min_value=1,
                    value=st.session_state.ocr_page_start,
                    key="page_start_input",
                    label_visibility="visible"
                )
            with col_x2:
                st.markdown("<br>", unsafe_allow_html=True)
                st.caption("ถึง N")
            st.session_state.ocr_page_start = start_page
            page_config = f"{start_page}-N"
        else:  # ระบุช่วง
            col_start, col_sep, col_end = st.columns([0.4, 0.2, 0.4])
            with col_start:
                start_page = st.number_input(
                    "หน้าเริ่มต้น:",
                    min_value=1,
                    value=st.session_state.ocr_page_start,
                    key="page_start_range",
                    label_visibility="visible"
                )
            with col_sep:
                st.markdown("<br>", unsafe_allow_html=True)
                st.caption("ถึง")
            with col_end:
                end_page = st.number_input(
                    "หน้าสุดท้าย:",
                    min_value=1,
                    value=max(st.session_state.ocr_page_start, 10),
                    key="page_end_range",
                    label_visibility="visible"
                )
            if start_page <= end_page:
                page_config = f"{start_page}-{end_page}"
            else:
                st.warning("⚠️ หน้าเริ่มต้น ≤ หน้าสุดท้าย")
                page_config = st.session_state.ocr_page_config
    
    st.session_state.ocr_page_config = page_config
    st.markdown("---")
    
    # Top Control Bar - Split into Left (Source) and Right (Output)
    col_left, col_right = st.columns([0.5, 0.5])
    
    # Left side - Source Files controls
    with col_left:
        # Add text input for manual path entry
        manual_path = st.text_input(
            "Source Folder Path:",
            value=st.session_state.ocr_source_folder if st.session_state.ocr_source_folder else r"D:\Project\ocr\source",
            help="Enter folder path manually or use 📁 button to browse",
            key="source_folder_input"
        )
        
        # Update session state if path is valid or create it
        if manual_path:
            # Create directory if it doesn't exist
            if not os.path.exists(manual_path):
                try:
                    os.makedirs(manual_path, exist_ok=True)
                    st.success(f"✅ Created folder: {manual_path}")
                except Exception as e:
                    st.warning(f"⚠️ Could not create source directory: {e}")
            
            if os.path.exists(manual_path) and os.path.isdir(manual_path):
                if st.session_state.ocr_source_folder != manual_path:
                    st.session_state.ocr_source_folder = manual_path
                    st.session_state.ocr_file_list_refresh += 1
            elif manual_path != st.session_state.ocr_source_folder:
                st.session_state.ocr_source_folder = None
        
        # PDF File Uploader - สำหรับ upload PDF ไปยัง source folder
        if st.session_state.ocr_source_folder and os.path.exists(st.session_state.ocr_source_folder):
            st.markdown("**📤 Upload PDF Files:**")
            uploaded_pdfs = st.file_uploader(
                "📄 Upload PDF Files to Source Folder",
                type=['pdf'],
                accept_multiple_files=True,
                help=f"Upload PDF files to: {st.session_state.ocr_source_folder}",
                key="pdf_uploader_ocr"
            )
            
            if uploaded_pdfs:
                saved_count = 0
                for uploaded_file in uploaded_pdfs:
                    try:
                        # Save to source folder
                        file_path = os.path.join(st.session_state.ocr_source_folder, uploaded_file.name)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        saved_count += 1
                    except Exception as e:
                        st.error(f"❌ Error saving {uploaded_file.name}: {e}")
                
                if saved_count > 0:
                    st.success(f"✅ Saved {saved_count} PDF file(s) to: {st.session_state.ocr_source_folder}")
                    st.session_state.ocr_file_list_refresh += 1
                    time.sleep(1)
                    st.rerun()
        
        col_btn1, col_btn2, col_btn3 = st.columns([0.2, 0.2, 0.2])
        
        with col_btn1:
            if st.button("📁", use_container_width=True, type="primary", help="Browse Folder (may not work in some environments)", key="select_folder_btn"):
                # Default initial directory
                default_source_dir = r"D:\Project\ocr\source"
                initial_dir = default_source_dir if os.path.exists(default_source_dir) else None
                selected_folder = select_folder_dialog(initial_dir=initial_dir)
                if selected_folder:
                    st.session_state.ocr_source_folder = selected_folder
                    st.session_state.ocr_file_list_refresh += 1
                    st.rerun()
        
        with col_btn2:
            if st.button("Run OCR", use_container_width=True, disabled=st.session_state.ocr_source_folder is None, help="Run OCR", key="run_ocr_btn"):
                if st.session_state.ocr_source_folder:
                    try:
                        source_path = st.session_state.ocr_source_folder
                        output_path = st.session_state.ocr_output_folder
                        page_config = st.session_state.ocr_page_config
                        
                        # ตรวจสอบว่าเลือก AI OCR แบบไหน
                        if st.session_state.ocr_type == "API Typhoon":
                            # เลือก script ตาม OS (Cross-platform)
                            if platform.system() == 'Windows':
                                script_file = os.path.join(os.getcwd(), "runocr.bat")
                                cmd = [script_file, source_path, output_path, page_config]
                            else:
                                script_file = os.path.join(os.getcwd(), "runocr.sh")
                                cmd = ["bash", script_file, source_path, output_path, page_config]
                            
                            if os.path.exists(script_file):
                                with st.spinner("Running API Typhoon OCR..."):
                                    result = subprocess.run(
                                        cmd,
                                        cwd=os.getcwd(),
                                        capture_output=True,
                                        text=True,
                                        shell=(platform.system() == 'Windows')
                                    )
                                
                                # แสดง output เพื่อ debug
                                if result.stdout:
                                    st.text("Output:")
                                    st.code(result.stdout, language="text")
                                
                                if result.returncode == 0:
                                    st.success("OCR process completed successfully!")
                                    st.session_state.ocr_file_list_refresh += 1
                                    st.rerun()
                                else:
                                    st.error(f"OCR process failed with return code: {result.returncode}")
                                    if result.stderr:
                                        st.error("Error details:")
                                        st.code(result.stderr, language="text")
                            else:
                                st.error(f"OCR script not found at: {script_file}")
                        else:
                            # เลือก script ตาม OS (Cross-platform) - Local OCR
                            if platform.system() == 'Windows':
                                script_file = os.path.join(os.getcwd(), "runocr_local.bat")
                                cmd = [script_file, source_path, output_path, page_config]
                            else:
                                script_file = os.path.join(os.getcwd(), "runocr_local.sh")
                                cmd = ["bash", script_file, source_path, output_path, page_config]
                            
                            if os.path.exists(script_file):
                                # แสดง spinner และรัน process
                                spinner_placeholder = st.empty()
                                with spinner_placeholder.container():
                                    with st.spinner("Running Local Typhoon OCR... This may take a while."):
                                        result = subprocess.run(
                                            cmd,
                                            cwd=os.getcwd(),
                                            capture_output=True,
                                            text=True,
                                            shell=(platform.system() == 'Windows')
                                        )
                                
                                # ลบ spinner หลังจาก process เสร็จ
                                spinner_placeholder.empty()
                                
                                # แสดง output ทั้งหมดเพื่อ debug
                                if result.stdout:
                                    st.text("Output:")
                                    st.code(result.stdout, language="text")
                                
                                if result.returncode == 0:
                                    st.success("OCR process completed successfully!")
                                    st.session_state.ocr_file_list_refresh += 1
                                    time.sleep(0.5)  # รอสักครู่ก่อน rerun
                                    st.rerun()
                                else:
                                    st.error(f"OCR process failed with return code: {result.returncode}")
                                    if result.stderr:
                                        st.error("Error details:")
                                        st.code(result.stderr, language="text")
                                    if result.stdout:
                                        st.warning("Check output above for details")
                            else:
                                st.error(f"OCR script not found at: {script_file}")
                    except Exception as e:
                        st.error(f"Error running OCR: {e}")
        
        with col_btn3:
            if st.button("🔄", use_container_width=True, help="Refresh"):
                st.session_state.ocr_file_list_refresh += 1
                st.rerun()
    
    # Right side - Output Files controls (aligned to right)
    with col_right:
        # Add text input for manual output path entry
        manual_output_path = st.text_input(
            "Output Folder Path:",
            value=st.session_state.ocr_output_folder if st.session_state.ocr_output_folder else DEFAULT_OUTPUT_PATH,
            help="Enter output folder path manually or use 📂 button to browse",
            key="output_folder_input"
        )
        
        # Update session state if path is valid
        if manual_output_path:
            if os.path.exists(manual_output_path) and os.path.isdir(manual_output_path):
                if st.session_state.ocr_output_folder != manual_output_path:
                    st.session_state.ocr_output_folder = manual_output_path
                    st.session_state.ocr_file_list_refresh += 1
            else:
                # Create directory if it doesn't exist
                try:
                    os.makedirs(manual_output_path, exist_ok=True)
                    if st.session_state.ocr_output_folder != manual_output_path:
                        st.session_state.ocr_output_folder = manual_output_path
                        st.session_state.ocr_file_list_refresh += 1
                except Exception as e:
                    st.warning(f"⚠️ Could not create output directory: {e}")
        
        col_spacer_right, col_btn4, col_btn5, col_btn6 = st.columns([0.5, 0.17, 0.17, 0.16])
        
        with col_btn4:
            if st.button("📂", use_container_width=True, help="Browse Output Folder (may not work in some environments)"):
                selected_output = select_folder_dialog(initial_dir=st.session_state.ocr_output_folder)
                if selected_output:
                    st.session_state.ocr_output_folder = selected_output
                    st.session_state.ocr_file_list_refresh += 1
                    st.rerun()
        
        with col_btn5:
            if st.button("🗑️", use_container_width=True, help="Delete All Output Files", type="secondary"):
                if st.session_state.ocr_output_folder and os.path.exists(st.session_state.ocr_output_folder):
                    output_files = get_files_in_folder(st.session_state.ocr_output_folder)
                    if output_files:
                        st.session_state.show_delete_confirm = True
                        st.rerun()
                    else:
                        st.info("No files found in output folder to delete")
                else:
                    st.warning("Output folder does not exist or not set")
        
        with col_btn6:
            if st.button("🔃", use_container_width=True, help="Refresh Output List"):
                st.session_state.ocr_file_list_refresh += 1
                st.rerun()
    
    # Confirmation Dialog - on right side (half screen)
    if st.session_state.show_delete_confirm:
        st.markdown("---")
        col_left_spacer, col_confirm_dialog = st.columns([0.5, 0.5])
        
        with col_left_spacer:
            pass  # Empty space on left
        
        with col_confirm_dialog:
            st.warning("⚠️ **Warning: You are about to delete all files in the output folder!**")
            
            col_confirm_cancel, col_confirm_ok = st.columns([0.5, 0.5])
            with col_confirm_cancel:
                if st.button("❌ Cancel", use_container_width=True, type="secondary"):
                    st.session_state.show_delete_confirm = False
                    st.rerun()
            with col_confirm_ok:
                if st.button("✅ OK", use_container_width=True, type="primary"):
                    st.session_state.show_delete_confirm = False
                    if st.session_state.ocr_output_folder and os.path.exists(st.session_state.ocr_output_folder):
                        output_files = get_files_in_folder(st.session_state.ocr_output_folder)
                        if output_files:
                            try:
                                deleted_count = 0
                                for file in output_files:
                                    file_path = os.path.join(st.session_state.ocr_output_folder, file)
                                    if os.path.isfile(file_path):
                                        os.remove(file_path)
                                        deleted_count += 1
                                if deleted_count > 0:
                                    st.success(f"✅ Deleted {deleted_count} file(s) from output folder")
                                    st.session_state.ocr_file_list_refresh += 1
                                    st.rerun()
                                else:
                                    st.info("No files to delete")
                            except Exception as e:
                                st.error(f"Error deleting files: {e}")
                                st.rerun()
    
    st.markdown("---")
    
    # Main Layout - Split View
    col_source, col_output = st.columns([0.5, 0.5], gap="medium")
    
    # Left Column - Source Files
    with col_source:
        if st.session_state.ocr_source_folder:
            source_files = get_files_in_folder(st.session_state.ocr_source_folder)
            if source_files:
                for file in source_files:
                    st.text(f"📄 {file}")
            else:
                st.info("No files found in source folder")
        else:
            st.info("👈 Please select a source folder")
    
    # Right Column - Output Files
    with col_output:
        # Output files list
        if os.path.exists(st.session_state.ocr_output_folder):
            output_files = get_files_in_folder(st.session_state.ocr_output_folder)
            if output_files:
                for file in output_files:
                    st.text(f"📄 {file}")
            else:
                st.info("No files found in output folder")
        else:
            st.warning(f"Output folder does not exist: {st.session_state.ocr_output_folder}")

# --- PAGE 2: Document Editor (Existing Feature) ---
def render_page_2():
    # Page selection at the top (compact)
    col_title, col_page = st.columns([0.75, 0.25])
    with col_title:
        st.title("📁 Document Editor")
    with col_page:
        st.markdown("<div style='padding-top: 1.5rem;'>", unsafe_allow_html=True)
        page_options = {
            "Page 1: AI OCR Dashboard": "🏠",
            "Page 2: Document Editor": "📄"
        }
        selected_page = st.selectbox(
            "Page:",
            options=list(page_options.keys()),
            index=0 if st.session_state.current_page == "Page 1: AI OCR Dashboard" else 1,
            format_func=lambda x: f"{page_options[x]} {x.split(':')[1].strip()}",
            label_visibility="visible"
        )
        st.markdown("</div>", unsafe_allow_html=True)
        if selected_page != st.session_state.current_page:
            st.session_state.current_page = selected_page
            st.rerun()
    
    col_list, col_viewer = st.columns([0.5, 0.5], gap="small")
    selected_file_path = None
    selected_page = 1

    with col_list:
        if st.session_state.df_data is None:
            # Path settings - เปิด path ได้โดยตรงบน Cloud/Server
            with st.expander("⚙️ ตั้งค่า Path (Advanced)", expanded=True):
                # กำหนด default path ตาม OS และ environment
                if platform.system() == 'Windows':
                    default_path = r"D:\Project\ocr\output"
                else:
                    # On Linux/Cloud: try common paths
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    # Try output folder in script directory first
                    output_in_script = os.path.join(script_dir, "output")
                    if os.path.exists(output_in_script):
                        default_path = output_in_script
                    elif os.path.exists("/mount/src/view_ocr/output"):
                        default_path = "/mount/src/view_ocr/output"  # Streamlit Cloud
                    else:
                        default_path = os.path.join(script_dir, "output")  # Use script directory
                
                if not os.path.exists(default_path): 
                    # Fallback to script directory
                    default_path = os.path.dirname(os.path.abspath(__file__))
                
                col_path, col_load = st.columns([0.85, 0.15])
                with col_path:
                    base_folder = st.text_input("Path หลัก:", value=st.session_state.get('doc_editor_path', default_path))
                    st.session_state.doc_editor_path = base_folder
                with col_load:
                    st.markdown("<div style='padding-top: 1.5rem;'></div>", unsafe_allow_html=True)
                    load_from_path = st.button("📂 Load", use_container_width=True, help="โหลดรายการไฟล์จาก path")
                
                # สร้างโฟลเดอร์ถ้ายังไม่มี
                if base_folder and not os.path.exists(base_folder):
                    try:
                        os.makedirs(base_folder, exist_ok=True)
                        st.success(f"✅ สร้างโฟลเดอร์แล้ว: {base_folder}")
                    except Exception as e:
                        st.error(f"❌ ไม่สามารถสร้างโฟลเดอร์ได้: {e}")
                
                # แสดงรายการไฟล์ Excel ใน folder
                if base_folder and os.path.exists(base_folder):
                    excel_files = [f for f in os.listdir(base_folder) if f.lower().endswith(('.xlsx', '.xls'))]
                    if excel_files:
                        st.markdown("**📁 ไฟล์ Excel ใน folder:**")
                        for idx, file_name in enumerate(sorted(excel_files)):
                            file_path = os.path.join(base_folder, file_name)
                            col_file, col_open = st.columns([0.8, 0.2])
                            with col_file:
                                st.text(f"📄 {file_name}")
                            with col_open:
                                if st.button("Open", key=f"open_file_{idx}", use_container_width=True):
                                    # โหลดไฟล์โดยตรงจาก path
                                    try:
                                        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', dtype=str)
                                        df = df.replace('nan', '')
                                        if "_chk" not in df.columns: 
                                            df.insert(0, "_chk", False)
                                        else: 
                                            df["_chk"] = df["_chk"].astype(bool)
                                        
                                        # Ensure columns exist
                                        if not find_column_name(df.columns, ["vendor", "code"]): 
                                            df["Vendor code"] = ""
                                        if not find_column_name(df.columns, ["vendor", "name"]): 
                                            df["Vendor Name"] = ""
                                        
                                        # เก็บ path ของไฟล์ที่เปิด
                                        st.session_state.df_data = df
                                        st.session_state.base_folder_cache = base_folder
                                        st.session_state.loaded_file_path = file_path
                                        st.session_state.current_sheet = pd.ExcelFile(file_path).sheet_names[0]
                                        st.session_state.data_version = 0
                                        load_vendor_master()
                                        st.success(f"✅ Loaded: {file_name}")
                                        time.sleep(0.5)
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ Error loading file: {e}")
                    else:
                        st.info("📭 ไม่พบไฟล์ Excel ใน folder นี้\n\n💡 **Tip:** Run OCR ก่อนเพื่อสร้างไฟล์ Excel หรือ Upload ไฟล์ Excel ผ่านปุ่มด้านล่าง")
                elif base_folder:
                    st.warning(f"⚠️ Path ไม่มีอยู่: {base_folder}\n\n💡 **Tip:** พิมพ์ path ที่ถูกต้อง เช่น `/mount/src/view_ocr` หรือ `/mount/src/view_ocr/output`")
            
            st.markdown("---")
            st.markdown("**หรือ Upload ไฟล์:**")
            
            # Use container for file uploader to ensure proper event handling
            upload_container = st.container()
            with upload_container:
                uploaded_file = st.file_uploader(
                    "📂 Drop Excel File Here", 
                    type=['xlsx'], 
                    label_visibility="visible",
                    accept_multiple_files=False,
                    key="excel_file_uploader",
                    help="Drag and drop an Excel file here or click Browse files"
                )
            
            if uploaded_file:
                st.session_state.uploaded_file_ref = uploaded_file
                st.session_state.current_sheet = get_sheet_names_fresh(uploaded_file)[0]
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, sheet_name=st.session_state.current_sheet, engine='openpyxl', dtype=str)
                df = df.replace('nan', '')
                if "_chk" not in df.columns: 
                    df.insert(0, "_chk", False)
                else: 
                    df["_chk"] = df["_chk"].astype(bool)
                
                # Ensure columns exist
                if not find_column_name(df.columns, ["vendor", "code"]): 
                    df["Vendor code"] = ""
                if not find_column_name(df.columns, ["vendor", "name"]): 
                    df["Vendor Name"] = ""

                st.session_state.df_data = df
                st.session_state.base_folder_cache = base_folder
                st.session_state.data_version = 0  # Reset version
                load_vendor_master()
                st.rerun()

        else:
            # Toolbar [Sheet] [Gen SAP] [Detail+Nav] [Reset] [Save] [Reload Vendor]
            c_sheet, c_gen, c_detail, c_reset, c_save, c_reload = st.columns([0.20, 0.15, 0.20, 0.12, 0.12, 0.12])
            with c_sheet:
                # รองรับทั้ง uploaded file และ loaded from path
                file_ref = st.session_state.uploaded_file_ref
                file_path = st.session_state.get('loaded_file_path')
                
                # ดึงรายชื่อ sheet
                if file_ref:
                    sheet_names = get_sheet_names_fresh(file_ref)
                elif file_path and os.path.exists(file_path):
                    sheet_names = pd.ExcelFile(file_path).sheet_names
                else:
                    sheet_names = [st.session_state.current_sheet] if st.session_state.current_sheet else ["Sheet1"]
                
                curr = st.session_state.current_sheet
                idx = sheet_names.index(curr) if curr in sheet_names else 0
                new_sheet = st.selectbox("Sheet", sheet_names, index=idx, label_visibility="collapsed")
                if new_sheet != curr:
                    st.session_state.current_sheet = new_sheet
                    # อ่านจาก file_ref หรือ file_path
                    if file_ref:
                        file_ref.seek(0)
                        df = pd.read_excel(file_ref, sheet_name=new_sheet, engine='openpyxl', dtype=str)
                    elif file_path and os.path.exists(file_path):
                        df = pd.read_excel(file_path, sheet_name=new_sheet, engine='openpyxl', dtype=str)
                    else:
                        st.error("ไม่พบไฟล์ที่จะอ่าน")
                        df = st.session_state.df_data
                    
                    df = df.replace('nan', '')
                    if "_chk" not in df.columns: 
                        df.insert(0, "_chk", False)
                    else: 
                        df["_chk"] = df["_chk"].astype(bool)
                    if not find_column_name(df.columns, ["vendor", "code"]): 
                        df["Vendor code"] = ""
                    if not find_column_name(df.columns, ["vendor", "name"]): 
                        df["Vendor Name"] = ""
                    st.session_state.df_data = df
                    st.session_state.data_version += 1  # Force refresh
                    st.rerun()

            with c_gen:
                if st.button("⚙️ Gen SAP", help="สร้างไฟล์ขึ้น SAP", use_container_width=True):
                    success, result_df = generate_sap_data(st.session_state.df_data)
                    if success:
                        # เรียกหน้าต่าง Save As ทันทีสำหรับไฟล์ SAP (บันทึกเป็น .txt)
                        fname = f"SAP_Import_{datetime.now().strftime('%Y%m%d')}.txt"
                        base_path = st.session_state.get('base_folder_cache', os.getcwd())
                        
                        # บันทึกเป็นไฟล์ .txt (ใช้ tab delimiter)
                        save_success, msg = save_txt_local(result_df, fname, base_path, delimiter='\t')
                        
                        if save_success:
                            st.toast(f"บันทึกไฟล์ SAP เรียบร้อย: {msg}", icon="✅")
                        elif msg != "Cancelled":
                            st.error(f"บันทึกไฟล์ SAP ไม่สำเร็จ: {msg}")
                    else:
                        st.error(f"สร้างข้อมูล SAP ไม่สำเร็จ: {result_df}")

            with c_detail:
                # --- Navigation Helper ---
                def nav(delta):
                    if st.session_state.df_data is not None and not st.session_state.df_data.empty:
                        curr = st.session_state.selected_row_idx if st.session_state.selected_row_idx is not None else 0
                        new_i = curr + delta
                        if 0 <= new_i < len(st.session_state.df_data):
                            st.session_state.selected_row_idx = new_i
                            st.session_state.df_data["_chk"] = False
                            st.session_state.df_data.at[new_i, "_chk"] = True
                            st.session_state.data_version += 1
                            st.rerun()

                c_main_btn, c_up, c_down = st.columns([0.6, 0.2, 0.2])
                
                if st.session_state.view_mode == 'list':
                    has_sel = not st.session_state.df_data[st.session_state.df_data["_chk"]].empty
                    with c_main_btn:
                        if st.button("📝 Detail", type="primary", disabled=not has_sel, use_container_width=True):
                            st.session_state.view_mode = 'detail'
                            st.rerun()
                else:
                    with c_main_btn:
                        if st.button("🔙 Back", help="กลับไปรายการ", use_container_width=True):
                            st.session_state.view_mode = 'list'
                            st.rerun()
                
                with c_up:
                    if st.button("🔼", help="Previous Record", use_container_width=True): 
                        nav(-1)
                with c_down:
                    if st.button("🔽️", help="Next Record", use_container_width=True): 
                        nav(1)

            with c_reset:
                if st.button("❌ New", help="เริ่มใหม่", use_container_width=True):
                    st.session_state.df_data = None
                    st.session_state.uploaded_file_ref = None
                    st.session_state.view_mode = 'list'
                    st.rerun()

            with c_save:
                if st.button("💾 Save As", help="บันทึกไฟล์ Excel", use_container_width=True):
                    if st.session_state.df_data is not None:
                        fname = st.session_state.uploaded_file_ref.name
                        base_path = st.session_state.get('base_folder_cache', os.getcwd())
                        success, msg = save_excel_local(st.session_state.df_data, fname, base_path)
                        if success: 
                            st.toast(f"บันทึกเรียบร้อย", icon="✅")
                            time.sleep(1)
                        elif msg != "Cancelled": 
                            st.error(f"บันทึกไม่สำเร็จ: {msg}")
            
            with c_reload:
                if st.button("🔄 Vendor", help="Reload Vendor Master", use_container_width=True):
                    master = load_vendor_master(force_reload=True)
                    if master is not None:
                        st.toast(f"✅ Reload Vendor Master สำเร็จ ({len(master)} rows)", icon="✅")
                    else:
                        st.error(f"❌ ไม่สามารถโหลด Vendor Master ได้ - Path: {VENDOR_MASTER_PATH}")

            # --- Content ---
            if st.session_state.view_mode == 'list':
                df_cols = st.session_state.df_data.columns
                # หา column name ที่ตรงกับ pattern (ลองหลาย pattern)
                # VendorID_OCR อาจเป็น "VendorID_OCR" หรือ "Vendor_OCR" หรือ "VendorID"
                col_vid = None
                for pattern in [["vendor", "id"], ["vendorid"], ["vendor", "ocr"]]:
                    col_vid = find_column_name(df_cols, pattern)
                    if col_vid:
                        break
                if not col_vid:
                    # ลองหาโดยตรง
                    for col in df_cols:
                        col_lower = str(col).lower()
                        if "vendorid" in col_lower or ("vendor" in col_lower and "id" in col_lower and "ocr" in col_lower):
                            col_vid = col
                            break
                col_vid = col_vid or "VendorID_OCR"
                
                # Branch_OCR อาจเป็น "Branch_OCR" หรือ "BranchOCR"
                col_branch = None
                for pattern in [["branch"]]:
                    col_branch = find_column_name(df_cols, pattern)
                    if col_branch:
                        break
                if not col_branch:
                    # ลองหาโดยตรง
                    for col in df_cols:
                        col_lower = str(col).lower()
                        if "branch" in col_lower and "ocr" in col_lower:
                            col_branch = col
                            break
                col_branch = col_branch or ("Branch_OCR" if "Branch_OCR" in df_cols else "BranchOCR")
                
                col_vcode = find_column_name(df_cols, ["vendor", "code"]) or "Vendor code"
                col_vname = find_column_name(df_cols, ["vendor", "name"]) or "Vendor Name"
                
                def on_editor_change():
                    current_key = f"main_editor_{st.session_state.data_version}"
                    if current_key not in st.session_state: 
                        return
                    
                    edited = st.session_state[current_key]["edited_rows"]
                    need_update_version = False
                    need_rerun = False
                    
                    # Debug: แสดง columns ที่เจอ
                    # st.info(f"DEBUG: col_vid={col_vid}, col_branch={col_branch}, col_vcode={col_vcode}, col_vname={col_vname}")
                    
                    for idx, changes in edited.items():
                        idx = int(idx)
                        
                        # Update ค่าทั้งหมดก่อน
                        for col, val in changes.items():
                            st.session_state.df_data.at[idx, col] = val
                        
                        # ตรวจสอบว่ามีการแก้ไข VendorID_OCR หรือ Branch_OCR หรือไม่
                        vendor_id_changed = col_vid in changes
                        branch_changed = col_branch in changes
                        
                        if vendor_id_changed or branch_changed:
                            # อ่านค่าปัจจุบันจาก df_data (หลังจาก update แล้ว)
                            cur_vendor = st.session_state.df_data.at[idx, col_vid]
                            cur_branch = st.session_state.df_data.at[idx, col_branch]
                            
                            # แปลงเป็น string และ strip
                            cur_vendor = str(cur_vendor).strip() if pd.notna(cur_vendor) and cur_vendor else ""
                            cur_branch = str(cur_branch).strip() if pd.notna(cur_branch) and cur_branch else ""
                            
                            # Debug: แสดงค่าที่จะ lookup
                            # st.warning(f"DEBUG: Looking up - VendorID={cur_vendor}, Branch={cur_branch}")
                            
                            # Lookup เฉพาะเมื่อทั้ง 2 field มีค่า
                            if cur_vendor and cur_branch:
                                info = lookup_vendor_info(cur_vendor, cur_branch, debug=False)
                                
                                if info and 'code' in info:
                                    if col_vcode in df_cols:
                                        st.session_state.df_data.at[idx, col_vcode] = info['code']
                                        need_update_version = True
                                        # st.success(f"DEBUG: Updated Vendor code to {info['code']}")
                                    if col_vname in df_cols and 'name' in info:
                                        st.session_state.df_data.at[idx, col_vname] = info['name']
                                        need_update_version = True

                    # Handle checkbox selection
                    newly_checked_idx = None
                    for idx, changes in edited.items():
                        if "_chk" in changes and changes["_chk"] is True:
                            newly_checked_idx = int(idx)
                            break
                    
                    if newly_checked_idx is not None:
                        st.session_state.df_data["_chk"] = False
                        st.session_state.df_data.at[newly_checked_idx, "_chk"] = True
                        need_update_version = True
                    
                    # Update data_version เพื่อ force refresh UI
                    # (ไม่ต้อง rerun เพราะ Streamlit จะ rerun อัตโนมัติหลัง callback)
                    if need_update_version:
                        st.session_state.data_version += 1
                
                edited_df = st.data_editor(
                    st.session_state.df_data,
                    key=f"main_editor_{st.session_state.data_version}",
                    on_change=on_editor_change,
                    use_container_width=True,
                    height=750,
                    hide_index=True,
                    column_config={"_chk": st.column_config.CheckboxColumn(label="✔", width="small")}
                )
                
                sel = edited_df[edited_df["_chk"]]
                st.session_state.selected_row_idx = sel.index[0] if not sel.empty else None

            elif st.session_state.view_mode == 'detail':
                st.markdown("---")
                
                # แสดงสถานะ highlight (ถ้ามี) และปุ่ม Clear
                if st.session_state.highlighted_field:
                    highlight_info = st.session_state.highlighted_field
                    num_positions = len(st.session_state.pdf_highlight_positions)
                    col_status, col_clear = st.columns([0.9, 0.1])
                    with col_status:
                        if num_positions > 0:
                            method = st.session_state.pdf_highlight_positions[0].get('method', 'unknown')
                            if method == 'pdf_text_layer':
                                st.success(f"🔍 Highlighting: **{highlight_info.get('field_name')}** = `{highlight_info.get('field_value')}` ({num_positions} exact position(s) found in PDF)")
                            else:
                                st.info(f"🔍 Highlighting: **{highlight_info.get('field_name')}** = `{highlight_info.get('field_value')}` ({num_positions} approximate position(s) from OCR text)")
                        else:
                            st.warning(f"⚠️ Highlighting: **{highlight_info.get('field_name')}** = `{highlight_info.get('field_value')}` (No matches found. Make sure OCR has been run and .txt file exists.)")
                    with col_clear:
                        if st.button("❌", key="clear_highlight", help="Clear highlight", use_container_width=True):
                            st.session_state.highlighted_field = None
                            st.session_state.pdf_highlight_positions = []
                            st.rerun()
                
                if st.session_state.selected_row_idx is not None:
                    current_idx = st.session_state.selected_row_idx
                    row_data = st.session_state.df_data.iloc[current_idx]
                    cols = [c for c in st.session_state.df_data.columns if c != "_chk"]
                    
                    df_cols = st.session_state.df_data.columns
                    col_vid = find_column_name(df_cols, ["vendor", "id"]) or "VendorID_OCR"
                    col_branch = find_column_name(df_cols, ["branch"]) or "BranchOCR"
                    col_vcode = find_column_name(df_cols, ["vendor", "code"]) or "Vendor code"
                    col_vname = find_column_name(df_cols, ["vendor", "name"]) or "Vendor Name"
                    
                    def update_val(col_name, idx):
                        input_key = f"det_{col_name}_{st.session_state.data_version}"
                        val = st.session_state[input_key]
                        
                        # Clean number format for InvAmtOCR columns (remove commas before saving)
                        if "InvAmtOCR" in str(col_name):
                            try:
                                # Remove commas and spaces
                                cleaned_val = str(val).replace(',', '').replace(' ', '').strip()
                                if cleaned_val and cleaned_val.lower() not in ['nan', 'none', '']:
                                    num_val = float(cleaned_val)
                                    # Save as number string without commas, preserve decimal places
                                    # If it's a whole number, save without .00, otherwise preserve decimals
                                    if num_val == int(num_val):
                                        val = str(int(num_val))
                                    else:
                                        # Keep up to 2 decimal places, remove trailing zeros
                                        val = f"{num_val:.2f}".rstrip('0').rstrip('.')
                            except (ValueError, TypeError):
                                pass  # If conversion fails, save as is
                        
                        st.session_state.df_data.at[idx, col_name] = val
                        
                        # ตรวจสอบว่ามีการแก้ไข VendorID_OCR หรือ Branch_OCR หรือไม่
                        if col_name in [col_vid, col_branch]:
                            # อ่านค่าปัจจุบันจาก df_data (หลังจาก update แล้ว)
                            cur_vendor = str(st.session_state.df_data.at[idx, col_vid]).strip() if pd.notna(st.session_state.df_data.at[idx, col_vid]) else ""
                            cur_branch = str(st.session_state.df_data.at[idx, col_branch]).strip() if pd.notna(st.session_state.df_data.at[idx, col_branch]) else ""
                            
                            # Lookup เฉพาะเมื่อทั้ง 2 field มีค่า
                            if cur_vendor and cur_branch:
                                info = lookup_vendor_info(cur_vendor, cur_branch, debug=False)
                                if info and 'code' in info:
                                    if col_vcode in df_cols:
                                        st.session_state.df_data.at[idx, col_vcode] = info['code']
                                    if col_vname in df_cols and 'name' in info:
                                        st.session_state.df_data.at[idx, col_vname] = info['name']
                                    # Update data_version เพื่อ force refresh UI
                                    # (ไม่ต้อง rerun เพราะ Streamlit จะ rerun อัตโนมัติหลัง callback)
                                    st.session_state.data_version += 1
                    
                    # ฟังก์ชันสำหรับ field focus และ highlight
                    def on_field_focus(col_name, field_value, idx, pdf_path, page_num):
                        """เมื่อ field ถูก focus ให้ค้นหาและ highlight ใน PDF"""
                        if field_value and str(field_value).strip() and field_value != 'nan' and pdf_path and os.path.exists(pdf_path):
                            # เก็บข้อมูล field ที่เลือก
                            clean_value = str(field_value).strip()
                            st.session_state.highlighted_field = {
                                'field_name': col_name,
                                'field_value': clean_value,
                                'row_idx': idx
                            }
                            
                            # ค้นหาตำแหน่งใน PDF
                            if os.path.splitext(pdf_path)[1].lower() == '.pdf':
                                positions = find_text_bbox_in_pdf(pdf_path, clean_value, page_num, field_name=col_name)
                                st.session_state.pdf_highlight_positions = positions
                                st.session_state.data_version += 1  # Force refresh
                                
                                # แสดง toast notification พร้อมข้อมูล debug
                                if positions:
                                    method = positions[0].get('method', 'unknown')
                                    matched_text = positions[0].get('text', '')
                                    match_score = positions[0].get('match_score', 0)
                                    confidence = positions[0].get('confidence', 0)
                                    
                                    if method == 'tesseract_ocr':
                                        st.toast(f"✅ Found: '{matched_text}' (Match: {match_score}%, Conf: {confidence:.0f}%)", icon="🔍")
                                        # แสดงข้อความเตือนถ้า matched text ไม่ตรงกับที่ค้นหา
                                        if matched_text.lower() != clean_value.lower():
                                            st.warning(f"⚠️ Found similar text: **{matched_text}** (searching for: {clean_value}). Position may be approximate.")
                                    elif method == 'pdf_text_layer':
                                        st.toast(f"✅ Found {len(positions)} match(es) in PDF text layer", icon="✅")
                                    else:
                                        st.toast(f"✅ Found {len(positions)} approximate position(s) from OCR text", icon="✅")
                                else:
                                    st.toast(f"⚠️ No matches found for '{col_name}'. Make sure OCR has been run.", icon="⚠️")

                    # เก็บ fpath และ pg ไว้ก่อน (จะหาในภายหลัง)
                    # แต่เราต้องหา fpath ก่อนเพื่อใช้ใน on_field_focus
                    # ให้เราย้ายการหา fpath มาด้านบน
                    temp_fpath = None
                    temp_pg = 1
                    if st.session_state.selected_row_idx is not None and st.session_state.df_data is not None:
                        temp_row = st.session_state.df_data.iloc[st.session_state.selected_row_idx]
                        temp_links = []
                        # รองรับทั้ง uploaded file และ loaded from path
                        temp_file_source = st.session_state.uploaded_file_ref
                        if temp_file_source is None:
                            temp_file_source = st.session_state.get('loaded_file_path')
                        if temp_file_source is not None and st.session_state.current_sheet is not None:
                            temp_links = extract_hyperlinks(temp_file_source, st.session_state.current_sheet)
                        if st.session_state.selected_row_idx < len(temp_links):
                            for k, v in temp_links[st.session_state.selected_row_idx].items():
                                # v is now a dict with 'target' and 'display' keys
                                if v and isinstance(v, dict):
                                    target = v.get('target', '')
                                    if target and ('.pdf' in target.lower() or '.png' in target.lower() or '.jpg' in target.lower() or '.jpeg' in target.lower()): 
                                        candidate_path = target.strip().replace('/', os.sep).replace('\\', os.sep)
                                        if not os.path.isabs(candidate_path):
                                            base = st.session_state.get('base_folder_cache', os.getcwd())
                                            candidate_path = os.path.join(base, candidate_path)
                                        candidate_path = os.path.normpath(candidate_path)
                                        if os.path.exists(candidate_path):
                                            temp_fpath = candidate_path
                                            break
                                        else:
                                            original_path = target.strip()
                                            if os.path.exists(original_path):
                                                temp_fpath = original_path
                                                break
                        
                        if not temp_fpath:
                            cols_lower = [str(c).lower() for c in st.session_state.df_data.columns]
                            fname = ""
                            if "filename" in cols_lower: 
                                fname = str(temp_row[st.session_state.df_data.columns[cols_lower.index("filename")]])
                            clean_name = re.sub(r'[\r\n\t"]', '', fname.replace("เปิดไฟล์", "").strip())
                            if clean_name:
                                base = st.session_state.get('base_folder_cache', os.getcwd())
                                temp_fpath = os.path.join(base, clean_name)
                                temp_fpath = os.path.normpath(temp_fpath)
                        
                        pg_cols = [c for c in st.session_state.df_data.columns if "page" in str(c).lower()]
                        if pg_cols:
                            try: 
                                temp_pg = int(float(temp_row[pg_cols[0]]))
                            except: 
                                temp_pg = 1

                    for i in range(0, len(cols), 2):
                        c1, c2 = st.columns(2)
                        col1 = cols[i]
                        with c1: 
                            # แบ่งเป็น 2 columns: field และปุ่ม highlight
                            field_col, btn_col = st.columns([0.85, 0.15])
                            
                            with field_col:
                                field_value = str(row_data[col1])
                                
                                # ถ้า column นี้มี hyperlink ให้แสดง display text แทน
                                if st.session_state.selected_row_idx < len(temp_links):
                                    link_info = temp_links[st.session_state.selected_row_idx].get(col1)
                                    if link_info and isinstance(link_info, dict):
                                        field_value = link_info.get('display', field_value)
                                
                                # Format date for InvDateOCR columns
                                field_value = format_date_value(field_value, col1)
                                # Format number for InvAmtOCR columns
                                field_value = format_number_value(field_value, col1)
                                
                                input_key = f"det_{col1}_{st.session_state.data_version}"
                                
                                new_value = st.text_input(
                                col1, 
                                    value=field_value, 
                                    key=input_key,
                                on_change=update_val,
                                    args=(col1, current_idx),
                                    label_visibility="visible"
                                )
                            
                            with btn_col:
                                # เพิ่ม padding-top เพื่อให้ปุ่มอยู่ระดับเดียวกับ text input
                                st.markdown('<div style="padding-top: 32px;"></div>', unsafe_allow_html=True)
                                
                                # ปุ่ม highlight
                                highlight_key = f"hl_{col1}_{st.session_state.data_version}"
                                if st.button("🔍", key=highlight_key, 
                                            help=f"Highlight '{col1}' in PDF", use_container_width=True):
                                    if new_value and str(new_value).strip() and new_value != 'nan':
                                        if temp_fpath and os.path.exists(temp_fpath):
                                            on_field_focus(col1, new_value, current_idx, temp_fpath, temp_pg)
                                            st.rerun()
                                        else:
                                            st.warning(f"⚠️ PDF file not found. Cannot highlight.")
                                    else:
                                        st.warning(f"⚠️ Field '{col1}' is empty. Cannot highlight.")
                            
                        if i+1 < len(cols):
                            col2 = cols[i+1]
                            with c2: 
                                field_col2, btn_col2 = st.columns([0.85, 0.15])
                                
                                with field_col2:
                                    field_value2 = str(row_data[col2])
                                    
                                    # ถ้า column นี้มี hyperlink ให้แสดง display text แทน
                                    if st.session_state.selected_row_idx < len(temp_links):
                                        link_info2 = temp_links[st.session_state.selected_row_idx].get(col2)
                                        if link_info2 and isinstance(link_info2, dict):
                                            field_value2 = link_info2.get('display', field_value2)
                                    
                                    # Format date for InvDateOCR columns
                                    field_value2 = format_date_value(field_value2, col2)
                                    # Format number for InvAmtOCR columns
                                    field_value2 = format_number_value(field_value2, col2)
                                    
                                    input_key2 = f"det_{col2}_{st.session_state.data_version}"
                                    
                                    new_value2 = st.text_input(
                                    col2, 
                                        value=field_value2,
                                        key=input_key2,
                                    on_change=update_val,
                                        args=(col2, current_idx),
                                        label_visibility="visible"
                                    )
                                
                                with btn_col2:
                                    # เพิ่ม padding-top เพื่อให้ปุ่มอยู่ระดับเดียวกับ text input
                                    st.markdown('<div style="padding-top: 32px;"></div>', unsafe_allow_html=True)
                                    
                                    highlight_key2 = f"hl_{col2}_{st.session_state.data_version}"
                                    if st.button("🔍", key=highlight_key2, 
                                                help=f"Highlight '{col2}' in PDF", use_container_width=True):
                                        if new_value2 and str(new_value2).strip() and new_value2 != 'nan':
                                            if temp_fpath and os.path.exists(temp_fpath):
                                                on_field_focus(col2, new_value2, current_idx, temp_fpath, temp_pg)
                                                st.rerun()
                                            else:
                                                st.warning(f"⚠️ PDF file not found. Cannot highlight.")
                                        else:
                                            st.warning(f"⚠️ Field '{col2}' is empty. Cannot highlight.")

    fpath = None
    pg = 1
    if st.session_state.selected_row_idx is not None and st.session_state.df_data is not None:
        row = st.session_state.df_data.iloc[st.session_state.selected_row_idx]
        
        # Method 1: Try to extract from hyperlinks first
        links = []
        # รองรับทั้ง uploaded file และ loaded from path
        file_source = st.session_state.uploaded_file_ref
        if file_source is None:
            file_source = st.session_state.get('loaded_file_path')
        
        if file_source is not None and st.session_state.current_sheet is not None:
            links = extract_hyperlinks(file_source, st.session_state.current_sheet)
            if st.session_state.selected_row_idx < len(links):
                for k, v in links[st.session_state.selected_row_idx].items():
                    # v is now a dict with 'target' and 'display' keys
                    if v and isinstance(v, dict):
                        target = v.get('target', '')
                        if target and ('.pdf' in target.lower() or '.png' in target.lower() or '.jpg' in target.lower() or '.jpeg' in target.lower()): 
                            candidate_path = target.strip()
                            # Normalize path separators for Windows
                            candidate_path = candidate_path.replace('/', os.sep).replace('\\', os.sep)
                            # If relative path, try to resolve it
                            if not os.path.isabs(candidate_path):
                                base = st.session_state.get('base_folder_cache', os.getcwd())
                                candidate_path = os.path.join(base, candidate_path)
                            # Normalize the path (resolve .. and .)
                            candidate_path = os.path.normpath(candidate_path)
                            # Check if file exists
                            if os.path.exists(candidate_path):
                                fpath = candidate_path
                                break
                            else:
                                # Try original path from hyperlink as-is (might be absolute)
                                original_path = target.strip()
                                if os.path.exists(original_path):
                                    fpath = original_path
                                    break
        
        # Method 2: Read directly from "Link PDF" column (even if it's not a hyperlink)
        if not fpath:
            cols_lower = [str(c).lower() for c in st.session_state.df_data.columns]
            
            # Try "Link PDF" column first
            link_pdf_col = None
            for col in st.session_state.df_data.columns:
                if "link" in str(col).lower() and "pdf" in str(col).lower():
                    link_pdf_col = col
                    break
            
            if link_pdf_col:
                link_value = row[link_pdf_col]
                if link_value and str(link_value).strip().lower() not in ['none', 'nan', '']:
                    # Extract path from HYPERLINK formula if present
                    link_str = str(link_value).strip()
                    
                    # Check if it's a HYPERLINK formula
                    if link_str.upper().startswith("=HYPERLINK"):
                        # Parse HYPERLINK formula: =HYPERLINK("target", "display")
                        matches = re.findall(r'["\']([^"\']+)["\']', link_str)
                        if len(matches) >= 1:
                            candidate_path = matches[0].strip()
                        else:
                            candidate_path = link_str
                    else:
                        # Direct path value
                        candidate_path = link_str
                    
                    # Normalize path
                    candidate_path = candidate_path.replace('/', os.sep).replace('\\', os.sep)
                    
                    # If relative path, try to resolve it
                    if not os.path.isabs(candidate_path):
                        base = st.session_state.get('base_folder_cache', os.getcwd())
                        candidate_path = os.path.join(base, candidate_path)
                    
                    candidate_path = os.path.normpath(candidate_path)
                    
                    # Check if file exists
                    if os.path.exists(candidate_path):
                        fpath = candidate_path
        
        # Method 3: Try "filename" column as fallback
        if not fpath:
            cols_lower = [str(c).lower() for c in st.session_state.df_data.columns]
            fname = ""
            if "filename" in cols_lower: 
                fname = str(row[st.session_state.df_data.columns[cols_lower.index("filename")]])
            
            if fname and str(fname).strip().lower() not in ['none', 'nan', '']:
                clean_name = re.sub(r'[\r\n\t"]', '', fname.replace("เปิดไฟล์", "").strip())
                if clean_name:
                    base = st.session_state.get('base_folder_cache', os.getcwd())
                    fpath = os.path.join(base, clean_name)
                    # Normalize the path
                    fpath = os.path.normpath(fpath)
        
        pg_cols = [c for c in st.session_state.df_data.columns if "page" in str(c).lower()]
        if pg_cols:
            try: 
                pg = int(float(row[pg_cols[0]]))
            except: 
                pg = 1

    with col_viewer:
        st.markdown('<div class="css-card" style="height: 100%;">', unsafe_allow_html=True) 
        if fpath:
            if os.path.exists(fpath):
                c_v_head, c_v_btn = st.columns([0.8, 0.2])
                with c_v_head: 
                    st.caption(f"📄 {os.path.basename(fpath)}")
                with c_v_btn:
                    if st.button("↗ Open", key="ext_open", use_container_width=True): 
                        open_file_external(fpath)
                ext = os.path.splitext(fpath)[1].lower()
                if ext == '.pdf': 
                    # Initialize zoom level in session state
                    if 'pdf_zoom_level' not in st.session_state:
                        st.session_state.pdf_zoom_level = 1.0
                    
                    # Zoom controls
                    col_zoom1, col_zoom2, col_zoom3, col_zoom4, col_zoom5 = st.columns([0.15, 0.15, 0.2, 0.15, 0.15])
                    with col_zoom1:
                        if st.button("🔍−", key="zoom_out", help="Zoom Out (-5%)", use_container_width=True):
                            st.session_state.pdf_zoom_level = max(0.25, st.session_state.pdf_zoom_level - 0.05)
                            st.rerun()
                    with col_zoom2:
                        if st.button("🔍+", key="zoom_in", help="Zoom In (+5%)", use_container_width=True):
                            st.session_state.pdf_zoom_level = min(3.0, st.session_state.pdf_zoom_level + 0.05)
                            st.rerun()
                    with col_zoom3:
                        zoom_display = f"{int(st.session_state.pdf_zoom_level * 100)}%"
                        st.markdown(f"<div style='text-align: center; padding-top: 0.5rem;'>{zoom_display}</div>", unsafe_allow_html=True)
                    with col_zoom4:
                        if st.button("↺ Reset", key="zoom_reset", help="Reset to 100%", use_container_width=True):
                            st.session_state.pdf_zoom_level = 1.0
                            st.rerun()
                    with col_zoom5:
                        # Zoom slider (ปรับละเอียด 5% ต่อครั้ง)
                        new_zoom = st.slider(
                            "Zoom",
                            min_value=0.25,
                            max_value=3.0,
                            value=st.session_state.pdf_zoom_level,
                            step=0.05,
                            key="zoom_slider",
                            label_visibility="collapsed"
                        )
                        if new_zoom != st.session_state.pdf_zoom_level:
                            st.session_state.pdf_zoom_level = new_zoom
                            st.rerun()
                    
                    # ตรวจสอบว่ามี highlight positions หรือไม่
                    highlight_pos = None
                    if st.session_state.highlighted_field and st.session_state.pdf_highlight_positions:
                        # ตรวจสอบว่าเป็น field เดียวกันและ row เดียวกัน
                        if (st.session_state.highlighted_field.get('row_idx') == st.session_state.selected_row_idx and
                            st.session_state.highlighted_field.get('field_value')):
                            highlight_pos = st.session_state.pdf_highlight_positions
                    
                    # Debug info และ Text Context (แสดงใน expander)
                    if st.session_state.highlighted_field:
                        with st.expander("🔍 Highlight Info & Text Context", expanded=False):
                            st.write("**Highlighted Field:**", st.session_state.highlighted_field)
                            st.write("**Positions Found:**", len(highlight_pos) if highlight_pos else 0)
                            
                            if highlight_pos:
                                st.write("**First 3 Positions:**")
                                st.json(highlight_pos[:3])
                                
                                # แสดงคำแนะนำ
                                if highlight_pos[0].get('method') == 'ocr_txt_approximate':
                                    st.info("💡 **Note:** This is an approximate position calculated from OCR text. The red border may not be exactly on the text but should be close to it.")
                                elif highlight_pos[0].get('method') == 'pdf_text_layer':
                                    st.success("✅ **Exact position** found in PDF text layer.")
                            
                            # แสดง text context จาก OCR text file
                            if highlight_pos and highlight_pos[0].get('method') != 'pdf_text_layer':
                                # สำหรับ scan PDF แสดง text context
                                try:
                                    pdf_dir = os.path.dirname(fpath)
                                    pdf_basename = os.path.splitext(os.path.basename(fpath))[0]
                                    txt_filename = f"{pdf_basename}_page{pg}.txt"
                                    
                                    possible_txt_paths = [
                                        os.path.join(pdf_dir, txt_filename),
                                        os.path.join(st.session_state.get('ocr_output_folder', DEFAULT_OUTPUT_PATH), txt_filename),
                                        os.path.join(DEFAULT_OUTPUT_PATH, txt_filename),
                                    ]
                                    
                                    txt_path = None
                                    for path in possible_txt_paths:
                                        if os.path.exists(path):
                                            txt_path = path
                                            break
                                    
                                    if txt_path:
                                        with open(txt_path, 'r', encoding='utf-8') as f:
                                            ocr_text = f.read()
                                        
                                        search_value = st.session_state.highlighted_field.get('field_value', '')
                                        if search_value:
                                            # หา text context รอบๆ text ที่พบ
                                            pattern = re.escape(search_value)
                                            matches = list(re.finditer(pattern, ocr_text, re.IGNORECASE))
                                            
                                            if matches:
                                                st.write("**Text Context from OCR:**")
                                                for i, match in enumerate(matches[:3]):  # แสดง 3 matches แรก
                                                    start = max(0, match.start() - 50)
                                                    end = min(len(ocr_text), match.end() + 50)
                                                    context = ocr_text[start:end]
                                                    highlighted_context = context.replace(
                                                        match.group(0), 
                                                        f"**{match.group(0)}**"
                                                    )
                                                    st.markdown(f"**Match {i+1}:** ...{highlighted_context}...")
                                            else:
                                                st.write("**Text Context:** Text not found in OCR file")
                                    else:
                                        st.write(f"**OCR Text File:** Not found ({txt_filename})")
                                except Exception as e:
                                    st.write(f"**Error reading OCR text:** {e}")
                    
                    render_pdf(fpath, pg, highlight_positions=highlight_pos, zoom_level=st.session_state.pdf_zoom_level)
                elif ext in ['.png', '.jpg', '.jpeg']: 
                    st.image(fpath, use_container_width=True)
                else: 
                    st.info(f"File format not supported: {ext}")
            else: 
                st.error(f"❌ File not found: {fpath}")
                # Show debug info in expander
                with st.expander("🔍 Debug Info", expanded=False):
                    st.write(f"Selected row index: {st.session_state.selected_row_idx}")
                    st.write(f"File path: {fpath}")
                    st.write(f"Base folder: {st.session_state.get('base_folder_cache', os.getcwd())}")
                    if st.session_state.selected_row_idx is not None and st.session_state.df_data is not None:
                        row = st.session_state.df_data.iloc[st.session_state.selected_row_idx]
                        st.write("Row data:", row.to_dict())
        else: 
            st.info("👈 Select a row to view document")
            if st.session_state.selected_row_idx is not None and st.session_state.df_data is not None:
                with st.expander("🔍 Debug Info", expanded=False):
                    st.write(f"Selected row index: {st.session_state.selected_row_idx}")
                    st.write("No file path found. Checking...")
                    row = st.session_state.df_data.iloc[st.session_state.selected_row_idx]
                    cols_lower = [str(c).lower() for c in st.session_state.df_data.columns]
                    st.write("Available columns:", list(st.session_state.df_data.columns))
                    if "filename" in cols_lower:
                        fname_col = st.session_state.df_data.columns[cols_lower.index("filename")]
                        st.write(f"Filename column value: {row[fname_col]}")
                    # รองรับทั้ง uploaded file และ loaded from path
                    debug_file_source = st.session_state.uploaded_file_ref
                    if debug_file_source is None:
                        debug_file_source = st.session_state.get('loaded_file_path')
                    if debug_file_source is not None:
                        links = extract_hyperlinks(debug_file_source, st.session_state.current_sheet)
                        if st.session_state.selected_row_idx < len(links):
                            st.write("Hyperlinks found:", links[st.session_state.selected_row_idx])
                        else:
                            st.write("Hyperlinks found: (empty or row index out of range)")
                    
                    # Show Link PDF column value
                    cols_lower = [str(c).lower() for c in st.session_state.df_data.columns]
                    link_pdf_col = None
                    for col in st.session_state.df_data.columns:
                        if "link" in str(col).lower() and "pdf" in str(col).lower():
                            link_pdf_col = col
                            break
                    
                    if link_pdf_col:
                        link_value = row[link_pdf_col]
                        st.write(f"**Link PDF column value:** `{link_value}`")
                        st.write(f"**Link PDF column type:** {type(link_value).__name__}")
                        
                        # Try to extract path
                        if link_value and str(link_value).strip().lower() not in ['none', 'nan', '']:
                            link_str = str(link_value).strip()
                            if link_str.upper().startswith("=HYPERLINK"):
                                matches = re.findall(r'["\']([^"\']+)["\']', link_str)
                                if len(matches) >= 1:
                                    extracted_path = matches[0]
                                    st.write(f"**Extracted path from formula:** `{extracted_path}`")
                                    # Try to resolve
                                    if not os.path.isabs(extracted_path):
                                        base = st.session_state.get('base_folder_cache', os.getcwd())
                                        resolved = os.path.join(base, extracted_path)
                                        resolved = os.path.normpath(resolved)
                                        st.write(f"**Resolved path:** `{resolved}`")
                                        st.write(f"**Path exists:** {os.path.exists(resolved)}")
        st.markdown('</div>', unsafe_allow_html=True)

# --- Main Router ---
if st.session_state.current_page == "Page 1: AI OCR Dashboard":
    render_page_1()
elif st.session_state.current_page == "Page 2: Document Editor":
    render_page_2()
